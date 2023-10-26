import datetime, os, shortuuid, calendar, threading
from openpyxl import load_workbook
from flask import request, render_template, current_app
from .cms_base import CmsFormViewBase
from models.domain_table import DomainTable, CardMerchantTable
from common_utils.utils_funcs import PagingCLS



class DomainAnalysis(CmsFormViewBase):
    title = '域名分析'
    show_menu = False
    add_url_rules = [['/domainAnalysis/', 'domainAnalysis']]
    template = 'domain/domainAnalysis.html'
    MCLS = DomainTable

    def get_recent_month(self, dt, months):
        month = dt.month - 1 + months
        year = dt.year + month // 12
        month = month % 12 + 1
        day = min(dt.day, calendar.monthrange(year, month)[1])
        return dt.replace(year=year, month=month, day=day)

    def view_get(self):
        self.context['title'] = self.title
        total = self.MCLS.count()
        guoqi = self.MCLS.find_many({'end_time': {'$lt': datetime.datetime.now()}})
        sanDate = self.get_recent_month(datetime.datetime.now(), 3)
        jijiang = self.MCLS.find_many({'end_time':{'$gt': datetime.datetime.now(), '$lt': sanDate}})
        zc_data = self.MCLS.find_many({'end_time': {'$gt': sanDate}})
        _data = {
            'guoqi_count': len(guoqi),
            'jijiang_count': len(jijiang),
            'zc_data_count': len(zc_data),
        }
        guoqi_bfb, jijiang_bfb, zc_data_bfb = 0, 0, 0
        if total:
            guoqi_bfb = round(len(guoqi) / total * 100, 2)
            jijiang_bfb = round(len(jijiang) / total * 100, 2)
            zc_data_bfb = round(len(zc_data) / total * 100, 2)
        _data['guoqi_bfb'] = guoqi_bfb
        _data['jijiang_bfb'] = jijiang_bfb
        _data['zc_data_bfb'] = zc_data_bfb

        self.context.update(_data)
        return render_template(self.template, **self.context)



class DomainManager(CmsFormViewBase):
    MCLS = DomainTable
    title = '域名管理'
    add_url_rules = [['/domainManager/', 'domainManager']]
    template = 'domain/domainManager.html'
    per_page = 50
    sort = [['is_top', -1], ['end_time', -1]]

    def get_recent_month(self, dt, months):
        month = dt.month - 1 + months
        year = dt.year + month // 12
        month = month % 12 + 1
        day = min(dt.day, calendar.monthrange(year, month)[1])
        return dt.replace(year=year, month=month, day=day)

    def get_edit_html(self):
        html = f'''
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>到期时间：</span></div>
                    <input type="text" class="form-control selectDateYMD" id="end_time" placeholder="到期时间" value="{ self.format_datetime(self.data_dict.get('end_time') or '', '%Y-%m-%d') }" readonly onmouseover="$.single_YY_MM_DD('.selectDateYMD')">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>备注：</span></div>
                    <input type="text" class="form-control" id="note" placeholder="备注" value="{self.data_dict.get('note') or ''}">
                </div>                
                <span class="btn btn-primary swal2-styled" onclick="post_data_edit('{self.data_uuid}')" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>        
            '''
        return self.xtjson.json_result(message=html)

    def fenXi_xlsx(self, filePath):
        datas = []
        wb = load_workbook(filePath)
        wa = wb.active
        for r in wa.rows:
            _d = {
                'uuid': shortuuid.uuid(),
                'account': str(list(r)[0].value or ''),
                'domain': str(list(r)[1].value or ''),
                'is_top': False,
            }
            try:
                delta = datetime.timedelta(days=int(list(r)[2].value))
                today = datetime.datetime.strptime('1899-12-30', '%Y-%m-%d') + delta
                _d['end_time'] = today
            except:
                try:
                    _d['end_time'] = list(r)[2].value
                except:
                    pass

            datas.append(_d)

        return datas[1:]

    def del_all(self):
        self.MCLS.delete_many({})

    def import_data(self, datas):
        for dv in datas:
            self.MCLS.update_one({'domain': dv.get('domain')}, {'$set': dv},upsert=True)

    def format_datetime(self, data, fmt=None):
        if isinstance(data, datetime.datetime):
            if fmt:
                return data.strftime(fmt)
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    def view_get(self):
        self.context['title'] = self.title
        self.context['format_datetime'] = self.format_datetime
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res

        filter_dict.update(res[0])
        context_res.update(res[1])
        domainStatu = request.args.get('statu', '')
        if domainStatu:
            context_res.update({'statu': domainStatu})
            sanDate = self.get_recent_month(datetime.datetime.now(), 3)
            if domainStatu == '1':
                filter_dict.update({'end_time': {'$gt': sanDate}})
                self.sort=[['end_time', 1]]

            if domainStatu == '2':
                filter_dict.update({'end_time': {'$gt': datetime.datetime.now(), '$lt': sanDate}})
            if domainStatu == '3':
                filter_dict.update({'end_time': {'$lt': datetime.datetime.now()}})

        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'importDomain':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('xlsx'):
                return self.xtjson.json_params_error('文件格式错误，只支持xlsx文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/importFile/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)

            datas = self.fenXi_xlsx(import_folder + new_filename + fext)
            if not datas:
                os.remove(import_folder + new_filename + fext)
                return self.xtjson.json_params_error('文件无数据！')

            threading.Thread(target=self.import_data, args=(datas,)).start()
            return self.xtjson.json_result(message='数据提交成功，后台储存中，请稍后刷新查看！')
        if self.action == 'del_all':
            threading.Thread(target=self.del_all).start()
            return self.xtjson.json_result(message='删除中， 请稍后刷新查看！')

    def post_data_other_way(self):
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        if self.action == 'edit_from_html':
            return self.get_edit_html()
        if self.action == 'edit_data':
            end_time = self.request_data.get('end_time')
            note = self.request_data.get('note') or ''
            if not end_time:
                return self.xtjson.json_params_error('到期时间为空！')
            try:
                end_time = datetime.datetime.strptime(end_time, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('到期时间格式错误！')
            self.data_dict['end_time'] = end_time
            self.data_dict['note'] = note
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()
        if self.action == 'update_top':
            if self.data_dict.get('is_top'):
                self.data_from['is_top'] = False
            else:
                self.data_from['is_top'] = True
            self.data_dict.update(self.data_from)
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()



class  CardMerchantManage(CmsFormViewBase):
    MCLS = CardMerchantTable
    title = '卡商管理'
    add_url_rules = [['/cardMerchant/', 'CardMerchant']]
    template = 'domain/cardMerchant.html'
    per_page = 50
    sort = [['is_top', -1], ['end_time', -1]]

    def get_recent_month(self, dt, months):
        month = dt.month - 1 + months
        year = dt.year + month // 12
        month = month % 12 + 1
        day = min(dt.day, calendar.monthrange(year, month)[1])
        return dt.replace(year=year, month=month, day=day)

    def get_edit_html(self):
        html = f'''
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>到期时间：</span></div>
                    <input type="text" class="form-control selectDateYMD" id="end_time" placeholder="到期时间" value="{ self.format_datetime(self.data_dict.get('end_time') or '', '%Y-%m-%d') }" readonly onmouseover="$.single_YY_MM_DD('.selectDateYMD')">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">后台备注：</span></div>
                    <input type="text" class="form-control" id="cmd_note" placeholder="后台备注" value="{self.data_dict.get('cmd_note') or ''}">
                </div>                
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">备注：</span></div>
                    <input type="text" class="form-control" id="note" placeholder="备注" value="{self.data_dict.get('note') or ''}">
                </div>                
                <span class="btn btn-primary swal2-styled" onclick="post_data_edit('{self.data_uuid}')" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>        
            '''
        return self.xtjson.json_result(message=html)

    def fenXi_xlsx(self, filePath):
        datas = []
        wb = load_workbook(filePath)
        wa = wb.active
        for r in wa.rows:
            _vd = []
            _v_statu = False
            for i in range(6):
                try:
                    _v = str(list(r)[i].value or '')
                except:
                    _v = ''
                if not _v_statu and _v:
                    _v_statu = True
                if _v.strip() == 'None':
                    _v = ''
                _vd.append(_v.strip())
            if not _v_statu:
                continue
            if '时间' in _vd[3]:
                continue

            _d = {
                'uuid': shortuuid.uuid(),
                'cardMerchant_name': _vd[0],
                'username': _vd[1],
                'cardNo': _vd[2],
                'department': _vd[4],
                'note': _vd[5],
                'is_top': False,
            }
            print('_vd:', _vd)

            try:
                if len(_vd) < 4 or not _vd[3]:
                    _d['end_time'] = None
                else:
                    if not _vd[3].isdigit() or '-' in _vd[3]:
                        return False, '时间格式错误!'
                    delta = datetime.timedelta(days=int(_vd[3]))
                    today = datetime.datetime.strptime('1899-12-30', '%Y-%m-%d') + delta
                    _d['end_time'] = today
            except Exception as e:
                return False, '时间格式错误!'

            datas.append(_d)

        return True, datas

    def del_all(self):
        self.MCLS.delete_many({})

    def import_data(self, datas):
        for dv in datas:
            self.MCLS.update_one({'username': dv.get('username')}, {'$set': dv},upsert=True)

    def format_datetime(self, data, fmt=None):
        if isinstance(data, datetime.datetime):
            if fmt:
                return data.strftime(fmt)
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    def view_get(self):
        self.context['title'] = self.title
        self.context['format_datetime'] = self.format_datetime
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res

        filter_dict.update(res[0])
        context_res.update(res[1])
        domainStatu = request.args.get('statu', '')
        if domainStatu:
            context_res.update({'statu': domainStatu})
            # sanDate = self.get_recent_month(datetime.datetime.now(), 3)
            sanDate = datetime.datetime.now()
            if domainStatu == '1':
                filter_dict.update({'end_time': {'$gt': sanDate}})
            if domainStatu == '2':
                filter_dict.update({'end_time': {'$gt': sanDate + datetime.timedelta(days = -31), '$lt': sanDate}})
            if domainStatu == '3':
                filter_dict.update({'end_time': {'$lt': sanDate + datetime.timedelta(days = -31)}})

        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'importDomain':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('xlsx'):
                return self.xtjson.json_params_error('文件格式错误，只支持xlsx文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/importFile/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)

            statu, datas = self.fenXi_xlsx(import_folder + new_filename + fext)
            if not statu:
                return self.xtjson.json_params_error(datas)
            if not datas:
                os.remove(import_folder + new_filename + fext)
                return self.xtjson.json_params_error('文件无数据！')

            threading.Thread(target=self.import_data, args=(datas,)).start()
            return self.xtjson.json_result(message='数据提交成功，后台储存中，请稍后刷新查看！')
        if self.action == 'del_all':
            threading.Thread(target=self.del_all).start()
            return self.xtjson.json_result(message='删除中， 请稍后刷新查看！')

    def post_data_other_way(self):
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        if self.action == 'edit_from_html':
            return self.get_edit_html()
        if self.action == 'edit_data':
            end_time = self.request_data.get('end_time')
            cmd_note = self.request_data.get('cmd_note') or ''
            note = self.request_data.get('note') or ''
            if not end_time:
                return self.xtjson.json_params_error('到期时间为空！')
            try:
                end_time = datetime.datetime.strptime(end_time, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('到期时间格式错误！')
            self.data_dict['end_time'] = end_time
            self.data_dict['cmd_note'] = cmd_note
            self.data_dict['note'] = note
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()
        if self.action == 'update_top':
            if self.data_dict.get('is_top'):
                self.data_from['is_top'] = False
            else:
                self.data_from['is_top'] = True
            self.data_dict.update(self.data_from)
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()


