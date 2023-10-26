import os
import random
import datetime
import threading
import shortuuid
from .cms_base import CmsFormViewBase
from flask import render_template, request, current_app
from models.fenxi_table import hy_sanfang_form_table, hy_caipiao_form_tabl, hy_zijing_form_table, agency_caipiao_form_table, agency_sanfang_form_table, agency_zijing_form_table, agency_touzhu_form_table, \
    game_yxtj_form_table, game_sfyx_form_table, platform_form_table, RechargeTable, UploadlogTable, agencyConfigTable
from common_utils.utils_funcs import PagingCLS
from openpyxl import load_workbook, Workbook
from constants import ASSETS_FOLDER, EXPORT_FOLDER, ExportStatu
from models.customer_table import ExportDataModel, CustomerTable
from common_utils.lqredis import SiteRedis



class DataFormViewBase(CmsFormViewBase):
    per_page = 30
    sort = [['form_date', -1]]

    def format_money(self, data, strftimeStr=''):
        if not data:
            return ''
        try:
            try:
                if strftimeStr:
                    return data.strftime(strftimeStr)
            except:
                pass
            return data.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return data

    def uploadFile(self):
        fileobj = request.files['upload']
        fname, fext = os.path.splitext(fileobj.filename)
        import_folder = current_app.static_folder + '/' + self.project_name + '/assets/upload/'
        if not os.path.exists(import_folder):
            os.makedirs(import_folder)

        new_filename = shortuuid.uuid()
        fileobj.save(import_folder + new_filename + fext)
        filePath = '/assets/upload/' + new_filename + fext
        return self.xtjson.json_result(message=filePath)

    def data_ch_func(self):
        _d = UploadlogTable.find_one({'table_code': self.MCLS.__tablename__}) or {}
        if not _d:
            return self.xtjson.json_params_error('数据已撤回！')

        if _d.get('statu'):
            return self.xtjson.json_params_error('数据已撤回！')

        for _dd in UploadlogTable.find_many({'table_code': self.MCLS.__tablename__}):
            if _dd.get('log_type') == 'add':
                self.MCLS.delete_one({'uuid': _dd.get('data_id')})
            if _dd.get('log_type') == 'update':
                _lowd = self.MCLS.find_one({'uuid': _dd.get('data_id')}) or {}
                if _lowd:
                    low_data = _dd.get('low_data') or {}
                    if low_data:
                        _lowd.update(low_data)
                        self.MCLS.save(_lowd)
            _dd['statu'] = True
            UploadlogTable.save(_dd)
        return self.xtjson.json_result()

    def uplaod_form_html(self):
        html = f'''
            <div class="input-group mb-3">
                <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>上传报表：</span></div>
                <input type="text" class="form-control " id="formPath" placeholder="上传报表">
                <div class="input-group-append file-button">
                    <span class="input-group-text" id="basic-addon2">上传</span>
                    <input type="file" id="upload1" onchange="upload_file_func($('#upload1'), $('#formPath'), 'uploadFile', '', '', '', 'progress')">
                </div>
            </div>      
            <!-- 添加上传文件的进度条 -->
            <div id="showbar" class="mb-3" style="display: none;">
                <div class="progress">
                    <div id="progressbar" class="progress-bar progress-bar-success" role="progressbar" aria-valuenow="0" aria-valuemin="0" aria-valuemax="100" style="min-width: 2em;width: 0%">0%</div>
                </div>           
            </div>                          
            <div class="input-group mb-3">
                <div class="input-group-prepend"><span class="input-group-text">上传日期：</span></div>
                <input type="text" class="form-control selectDateYMD" onmouseenter="$.single_YY_MM_DD('.selectDateYMD')" value="{ (datetime.datetime.now() + datetime.timedelta(days=-1)).strftime('%Y-%m-%d') }" id="form_date" placeholder="上传日期">
            </div>
            <span class="btn btn-primary swal2-styled" onclick="uploadForm_func()" style="margin: 15px 10px 0px;">确定</span>
            <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 15px 10px 0px;">取消</span>
        '''
        return self.xtjson.json_result(message=html)

    def dataform_other_way(self):
        pass

    def del_search_data(self):
        filter_dict = {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if statu:
            filter_dict.update(res[0])
        if not filter_dict:
            return self.xtjson.json_params_error('请先搜索数据！')
        self.MCLS.delete_many(filter_dict)
        return self.xtjson.json_result()

    def view_get(self):
        self.context['title'] = self.title
        self.context['title'] = self.title
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['format_money'] = self.format_money
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'upload_form_html':
            return self.uplaod_form_html()
        if self.action == 'uploadFile':
            return self.uploadFile()
        if self.action == 'del_all':
            self.MCLS.delete_many({})
            return self.xtjson.json_result()
        if self.action == 'del_search_data':
            return self.del_search_data()
        if self.action == 'data_ch':
            return self.data_ch_func()
        res = self.dataform_other_way()
        if res:
            return res



class CustomerCpFormView(DataFormViewBase):
    title = '会员-彩票报表'
    show_menu = False
    add_url_rules = [['/customer/CpForm', 'customer_CpForm']]
    template = 'data_form/CpForm.html'
    MCLS = hy_caipiao_form_tabl
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '账号' not in one_row_data or '上级代理' not in one_row_data or '代理赔率（金额）' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            _upload_log = {
                'table_code': self.MCLS.__tablename__,
                'statu': False,
                'file_path': formPath,
            }
            _upload_log_uuid = UploadlogTable.insert_one(_upload_log)

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'account': _vv[0],
                    'name': _vv[1],
                    'parent_agent': _vv[2],
                    'betting_count': int(_vv[3]) or 0,
                    'betting_money': float(_vv[4]) or 0,
                    'profit_money': float(_vv[5]) or 0,
                    'agent_odds': float(_vv[6]) or 0,
                    'agent_backwater': float(_vv[7]) or 0,
                    'vip_winlose_money': float(_vv[8]) or 0,
                    'actual_backwater': float(_vv[9]) or 0,
                    'actual_winlose_money': float(_vv[10]) or 0,
                    'form_date': form_date,
                    'upload_log_id': _upload_log_uuid,
                }
                if self.MCLS.find_one({'account': _d.get('account'), 'name': _d.get('name'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            if datas:
                UploadlogTable.delete_many({'table_code': self.MCLS.__tablename__})
            for dd in datas:
                _up_log = {
                    'table_code': self.MCLS.__tablename__,
                    'statu': False,
                }
                _low_d = self.MCLS.find_one({'account': dd.get('account'), 'name': dd.get('name'), 'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    low_data = {}
                    low_data.update(_low_d)
                    if '_id' in low_data:
                        low_data.pop('_id')
                    if 'form_date' in low_data:
                        low_data.pop('form_date')
                    _up_log.update({
                        'data_id': _low_d.get('uuid'),
                        'log_type': 'update',
                        'low_data': low_data
                    })
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    data_id = self.MCLS.insert_one(dd)
                    _up_log.update({
                        'data_id': data_id,
                        'log_type': 'add',
                        'low_data': {}
                    })
                UploadlogTable.insert_one(_up_log)
            return self.xtjson.json_result()



class CustomerSfyxFormView(DataFormViewBase):
    title = '会员-第三方游戏报表'
    show_menu = False
    add_url_rules = [['/customer/SfyxForm', 'customer_SfyxForm']]
    template = 'data_form/SfyxForm.html'
    MCLS = hy_sanfang_form_table
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '有效投注金额' not in one_row_data or '投注金额' not in one_row_data or '会员返水金额' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'account': _vv[0],
                    'name': _vv[1],
                    'parent_agent': _vv[2],
                    'bs_number': int(_vv[3]) or 0,
                    'betting_money': float(_vv[4]) or 0,
                    'yx_betting_money': float(_vv[5]) or 0,
                    'sf_jackpot_yc_money': float(_vv[6]) or 0,
                    'sf_jackpot_pj_money': float(_vv[7]) or 0,
                    'sy_money': float(_vv[8]) or 0,
                    'fs_money': float(_vv[9]) or 0,
                    # 'sy_fs_money': float(_vv[10]) or 0,
                    'actual_winlose_money': float(_vv[10]) or 0,
                    'other_money': float(_vv[11]) or 0,
                    'form_date': form_date
                }
                if self.MCLS.find_one({'account': _d.get('account'), 'name': _d.get('name'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            for dd in datas:
                _low_d = self.MCLS.find_one({'account': dd.get('account'), 'name': dd.get('name'), 'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    self.MCLS.insert_one(dd)
            return self.xtjson.json_result()



class CustomerZjFormView(DataFormViewBase):
    title = '会员-资金报表'
    show_menu = False
    add_url_rules = [['/customer/ZjForm', 'customer_ZjForm']]
    template = 'data_form/ZjForm.html'
    MCLS = hy_zijing_form_table
    per_page = 30

    def fx_func(self, form_date1, form_date2, uuid, export_folder, filename, datas1):
        export_data = ExportDataModel.find_one({'uuid': uuid}) or {}
        end_time2 = datetime.datetime(form_date2.year, form_date2.month, form_date2.day, 23, 59, 59)
        resut_datas = []
        for data in datas1:
            cdata2 = self.MCLS.find_one({'form_date': {'$gte': form_date2, '$lte': end_time2}})
            _v1 = data.get('chongzhi_money') or 0
            _v2 = cdata2.get('chongzhi_money') or 0
            cz = round((_v1 - _v2), 2)
            if cz <= 0:continue
            _ccc = CustomerTable.find_one({'account': data.get('account')}) or {}
            resut_datas.append({
                'account': data.get('account'),
                'cz': cz,
                'telephone': _ccc.get('user_tele') or '',
            })

        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['账户', '金额差', '手机号']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            for data in resut_datas:
                row += 1
                wa.cell(row=row, column=1, value=str(data.get('account') or ''))
                wa.cell(row=row, column=2, value=str(self.format_money(data.get('cz') or '0')))
                wa.cell(row=row, column=3, value=str(data.get('telephone') or ''))
                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def get_duibi_fenxi_html(self):
        html = f'''                    
            <div class="input-group mb-3">
                <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>分析日期：</span></div>
                <input type="text" class="form-control selectDateYMD" onmouseenter="$.single_YY_MM_DD('.selectDateYMD')" value="" id="form_date1" placeholder=分析日期">
            </div>
            <div class="input-group mb-3">
                <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>对比日期：</span></div>
                <input type="text" class="form-control selectDateYMD" onmouseenter="$.single_YY_MM_DD('.selectDateYMD')" value="" id="form_date2" placeholder="对比日期">
            </div>
            <span class="btn btn-primary swal2-styled" onclick="fx_func()" style="margin: 15px 10px 0px;">确定</span>
            <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 15px 10px 0px;">取消</span>        
        '''
        return self.xtjson.json_result(message=html)

    def uploadForm_func(self):
        fugai = self.request_data.get('fugai')
        formPath = self.request_data.get('formPath')
        form_date_low = self.request_data.get('form_date')
        if not formPath or not form_date_low:
            return self.xtjson.json_params_error()
        import_folder = current_app.static_folder + '/' + self.project_name + formPath
        if not os.path.exists(import_folder):
            return self.xtjson.json_params_error('文件不存在！')
        try:
            form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
        except:
            return self.xtjson.json_params_error('日期错误！')

        try:
            dataf = open(import_folder, 'r', encoding='utf8').read()
        except:
            try:
                dataf = open(import_folder, 'r', encoding='gbk').read()
            except:
                return self.xtjson.json_params_error('读取文件失败！')

        _datas = dataf.split('\n')
        one_row_data = _datas[0]
        if '账号' not in one_row_data or '线上充值' not in one_row_data or '线下充值' not in one_row_data or '后台扣钱' not in one_row_data:
            return self.xtjson.json_params_error('内容错误！')

        start_time = form_date
        end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
        datas = []
        for dd in _datas[1:]:
            _v = dd.replace('"', '').replace('=', '').strip()
            if not _v:
                continue
            _vv = _v.split(',')
            if not _vv or not _vv[0]:
                continue
            _d = {
                'account': _vv[0],
                'parent_agent': _vv[1],
                'chongzhi_money': float(_vv[2]) or 0,
                'xscz_money': float(_vv[3]) or 0,
                'xxcz_money': float(_vv[4]) or 0,
                'hdjq_money': float(_vv[5]) or 0,
                'tixian_money': float(_vv[6]) or 0,
                'hdkq_money': float(_vv[7]) or 0,
                'hd_money': float(_vv[8]) or 0,
                'hb_money': float(_vv[9]) or 0,
                'czyh_money': float(_vv[10]) or 0,
                'form_date': form_date
            }
            if self.MCLS.find_one({'account': _d.get('account'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                if not fugai or fugai == 'false':
                    _result = {
                        'form_date': form_date_low,
                        'formPath': formPath,
                        'cfState': True,
                    }
                    return self.xtjson.json_result(data=_result)

            total = 0
            for f in ['chongzhi_money', 'xscz_money', 'xxcz_money', 'hdjq_money', 'tixian_money', 'hdkq_money',
                      'hd_money', 'hb_money', 'czyh_money']:
                total += float(_d.get(f) or 0)
            if total <= 0:
                continue
            datas.append(_d)
        for dd in datas:
            _low_d = self.MCLS.find_one(
                {'account': dd.get('account'), 'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
            if _low_d:
                _low_d.update(dd)
                self.MCLS.save(_low_d)
            else:
                self.MCLS.insert_one(dd)
        return self.xtjson.json_result()

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            return self.uploadForm_func()
        if self.action == 'get_duibi_fenxi_html':
            return self.get_duibi_fenxi_html()
        if self.action == 'fz_data':
            form_date1 = self.request_data.get('form_date1')
            form_date2 = self.request_data.get('form_date2')
            if not form_date1 or not form_date2:
                return self.xtjson.json_params_error()

            try:
                form_date1 = datetime.datetime.strptime(form_date1, '%Y-%m-%d')
                form_date2 = datetime.datetime.strptime(form_date2, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error()

            end_time1 = datetime.datetime(form_date1.year, form_date1.month, form_date1.day, 23, 59, 59)
            datas1 = self.MCLS.find_many({'form_date': {'$gte': form_date1, '$lte': end_time1}})

            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            export_data = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas1),
                'out_count': 0,
                'note': '会员报表-资金报表-分析' + datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            }
            uuid = ExportDataModel.insert_one(export_data)

            threading.Thread(target=self.fx_func, args=(form_date1, form_date2, uuid, export_folder, filename, datas1)).start()
            return self.xtjson.json_result()



class AgencyCpFormView(DataFormViewBase):
    title = '代理-彩票报表'
    show_menu = False
    add_url_rules = [['/agency/CpForm', 'agency_CpForm']]
    template = 'data_form/agency_CpForm.html'
    MCLS = agency_caipiao_form_table
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '账号' not in one_row_data or '下注会员数' not in one_row_data or '投注金额' not in one_row_data or '实际退水' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'account': _vv[0],
                    'name': _vv[1],
                    'xzhy_count': int(_vv[2]) or 0,
                    'betting_count': int(_vv[3]) or 0,
                    'betting_money': float(_vv[4]) or 0,
                    'profit_money': float(_vv[5]) or 0,
                    'agent_odds': float(_vv[6]) or 0,
                    'agent_backwater': float(_vv[7]) or 0,
                    'vip_winlose_money': float(_vv[8]) or 0,
                    'actual_backwater': float(_vv[9]) or 0,
                    'actual_winlose_money': float(_vv[10]) or 0,
                    'form_date': form_date
                }
                if self.MCLS.find_one({'account': _d.get('account'), 'name': _d.get('name'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            for dd in datas:
                _low_d = self.MCLS.find_one({'account': dd.get('account'), 'name': dd.get('name'),  'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    self.MCLS.insert_one(dd)
            return self.xtjson.json_result()



class AgencySfyxFormView(DataFormViewBase):
    title = '代理-第三方游戏报表'
    show_menu = False
    add_url_rules = [['/agency/SfyxForm', 'agency_SfyxForm']]
    template = 'data_form/agency_SfyxForm.html'
    MCLS = agency_sanfang_form_table
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '账号' not in one_row_data or '会员数' not in one_row_data or '有效投注金额' not in one_row_data or '代理返水金额' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'account': _vv[0],
                    'name': _vv[1],
                    'parent_agent': _vv[2],
                    'hy_count': int(_vv[3]) or 0,
                    'bs_number': int(_vv[4]) or 0,
                    'betting_money': float(_vv[5]) or 0,
                    'yx_betting_money': float(_vv[6]) or 0,
                    'sf_jackpot_yc_money': float(_vv[7]) or 0,
                    'sf_jackpot_pj_money': float(_vv[8]) or 0,
                    'sy_money': float(_vv[9]) or 0,
                    'hyfs_money': float(_vv[10]) or 0,
                    'dlfs_money': float(_vv[11]) or 0,
                    'sjsy_money': float(_vv[12]) or 0,
                    'form_date': form_date
                }
                if self.MCLS.find_one({'account': _d.get('account'), 'name': _d.get('name'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            for dd in datas:
                _low_d = self.MCLS.find_one({'account': dd.get('account'), 'name': dd.get('name'),  'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    self.MCLS.insert_one(dd)
            return self.xtjson.json_result()



class AgencyZjFormView(DataFormViewBase):
    title = '代理-资金报表'
    show_menu = False
    add_url_rules = [['/agency/ZjForm', 'agency_ZjForm']]
    template = 'data_form/agency_ZjForm.html'
    MCLS = agency_zijing_form_table
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '账号' not in one_row_data or '充值金额' not in one_row_data or '后台加钱' not in one_row_data or '后台扣钱' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'account': _vv[0],
                    'chongzhi_money': float(_vv[1]) or 0,
                    'hdjq_money': float(_vv[2]) or 0,
                    'tixian_money': float(_vv[3]) or 0,
                    'hdkq_money': float(_vv[4]) or 0,
                    'hd_money': float(_vv[5]) or 0,
                    'hb_money': float(_vv[6]) or 0,
                    'czyh_money': float(_vv[7]) or 0,
                    'form_date': form_date
                }
                if self.MCLS.find_one({'account': _d.get('account'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)

                total = 0
                for f in ['chongzhi_money', 'hdjq_money', 'tixian_money', 'hdkq_money', 'hd_money', 'hb_money', 'czyh_money']:
                    total += float(_d.get(f) or 0)
                if total <= 0:
                    continue
                datas.append(_d)
            for dd in datas:
                _low_d = self.MCLS.find_one({'account': dd.get('account'),  'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    self.MCLS.insert_one(dd)
            return self.xtjson.json_result()



class AgencyTzFormView(DataFormViewBase):
    title = '代理-资金报表'
    show_menu = False
    add_url_rules = [['/agency/TzForm', 'agency_TzForm']]
    template = 'data_form/agency_TzForm.html'
    MCLS = agency_touzhu_form_table
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '账号' not in one_row_data or '用户数' not in one_row_data or '投注笔数' not in one_row_data or '有效投注金额' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'account': _vv[0],
                    'customer_count': int(_vv[1]) or 0,
                    'tzbs_count': int(_vv[2]) or 0,
                    'betting_money': float(_vv[3]) or 0,
                    'yx_betting_money': float(_vv[4]) or 0,
                    'sy_money': float(_vv[5]) or 0,
                    'form_date': form_date
                }
                if self.MCLS.find_one({'account': _d.get('account'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            for dd in datas:
                _low_d = self.MCLS.find_one({'account': dd.get('account'),  'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    self.MCLS.insert_one(dd)
            return self.xtjson.json_result()



class GameYXTJFormView(DataFormViewBase):
    title = '游戏-游戏统计报表'
    show_menu = False
    add_url_rules = [['/game/YXTJ', 'game_yxtj']]
    template = 'data_form/game_yxtj.html'
    MCLS = game_yxtj_form_table
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '游戏' not in one_row_data or '盈利投注金额' not in one_row_data or '代理赔率金额' not in one_row_data or '代理返水金额' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'game_name': _vv[0],
                    'bs_number': int(_vv[1]) or 0,
                    'betting_money': float(_vv[2]) or 0,
                    'profit_money': float(_vv[3]) or 0,
                    'agent_odds': float(_vv[4]) or 0,
                    'agent_backwater': float(_vv[5]) or 0,
                    'vip_winlose_money': float(_vv[6]) or 0,
                    'actual_backwater': float(_vv[7]) or 0,
                    'actual_winlose_money': float(_vv[8]) or 0,
                    'xzhy_count': float(_vv[9]) or 0,
                    'form_date': form_date
                }
                if self.MCLS.find_one({'game_name': _d.get('game_name'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            if datas:
                UploadlogTable.delete_many({'table_code': self.MCLS.__tablename__})
            for dd in datas:
                _up_log = {
                    'table_code': self.MCLS.__tablename__,
                    'statu': False,
                }
                _low_d = self.MCLS.find_one({'game_name': dd.get('game_name'),  'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    low_data = {}
                    low_data.update(_low_d)
                    if '_id' in low_data:
                        low_data.pop('_id')
                    if 'form_date' in low_data:
                        low_data.pop('form_date')
                    _up_log.update({
                        'data_id': _low_d.get('uuid'),
                        'log_type': 'update',
                        'low_data': low_data
                    })
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    data_id = self.MCLS.insert_one(dd)
                    _up_log.update({
                        'data_id': data_id,
                        'log_type': 'add',
                        'low_data': {}
                    })
                UploadlogTable.insert_one(_up_log)
            return self.xtjson.json_result()



class GameSFYXFormView(DataFormViewBase):
    title = '游戏-第三方游戏统计'
    show_menu = False
    add_url_rules = [['/game/SFYX', 'game_sfyx']]
    template = 'data_form/game_sfyx.html'
    MCLS = game_sfyx_form_table
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')

            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('\n')
            one_row_data = _datas[0]
            if '平台' not in one_row_data or '有效投注金额' not in one_row_data or '第三方Jackpot预抽金额' not in one_row_data or '第三方Jackpot派彩金额' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'platform': _vv[0],
                    'yx_betting_money': float(_vv[1]) or 0,
                    'sf_jackpot_yc_money': float(_vv[2]) or 0,
                    'sf_jackpot_pj_money': float(_vv[3]) or 0,
                    'sy_money': float(_vv[4]) or 0,
                    'fs_money': float(_vv[5]) or 0,
                    'sjsy_money': float(_vv[6]) or 0,
                    'form_date': form_date
                }
                if self.MCLS.find_one({'platform': _d.get('platform'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            for dd in datas:
                _low_d = self.MCLS.find_one({'platform': dd.get('platform'),  'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    self.MCLS.insert_one(dd)
            return self.xtjson.json_result()



class PlatformFormView(DataFormViewBase):
    title = '平台报表'
    show_menu = False
    add_url_rules = [['/platformForm', 'platformForm']]
    template = 'data_form/platformForm.html'
    MCLS = platform_form_table
    per_page = 30

    def add_form_data_func(self):
        data_text = self.request_data.get('data_text')
        if not data_text or not data_text.strip():
            return self.xtjson.json_params_error('操作失败！')

        jx_datas = []
        try:
            data_text = data_text.replace('\r\n', '\n')
            for dl in data_text.split('\n\n'):
                if not dl or not dl.strip():
                    continue
                jx_datas.append(dl.strip().replace('\n', '').replace(' ', '').replace(',', '').replace('\t', ''))
        except:
            return self.xtjson.json_params_error('数据分析失败！')

        if len(jx_datas) != 15:
            return self.xtjson.json_params_error('数据解析错误！')

        _data = {}
        statuerror = False
        for index, dd in enumerate(jx_datas):
            if index == 0:
                try:
                    form_date = datetime.datetime.strptime(dd, '%Y-%m-%d')
                except:
                    statuerror = True
                    break
                _data.update({
                    'form_date': form_date,
                })
            if index == 1:
                _v = dd.split('/')
                if len(_v) != 3:
                    statuerror = True
                    break
                _data.update({
                    'data_date': _v[0],
                    'app_regist': _v[1],
                    'shou_chong': _v[2],
                })
            if index == 2:
                _v = dd.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'xzyxhy_cp': _v[0],
                    'xzyxhy_dsfyx': _v[1],
                })
            if index == 3:
                _v = dd.split('/')
                if len(_v) != 1:
                    statuerror = True
                    break
                _data.update({
                    'yxyh_count': int(_v[0]),
                })
            if index == 4:
                _v = dd.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'yxyh_cp_count': _v[0],
                    'yxyh_sfyx_count': _v[1],
                })
            if index == 5:
                _v = dd.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'cpye_money': float(_v[0]),
                    'sfye_money': float(_v[1]),
                })
            if index == 6:
                _v = dd.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'recharge_money': float(_v[0]),
                    'withdraw_money': float(_v[1]),
                })
            if index == 7:
                _v = dd.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'touzhu_money': float(_v[0]),
                    'tsze_money': float(_v[1]),
                })
            if index == 8:
                _v = dd.split('/')
                if len(_v) != 1:
                    statuerror = True
                    break
                _data.update({
                    'sjsy_money': float(_v[0]),
                })
            if index == 9:
                _v = dd.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'agent_odds': float(_v[0]),
                    'tuishui_money': float(_v[1]),
                })
            if index == 10:
                _v = dd.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'jackpot': float(_v[0]),
                    'jackBonus': float(_v[1]),
                })
            if index == 11:
                _v = dd.split('/')
                if len(_v) != 3:
                    statuerror = True
                    break
                _data.update({
                    'sfyx_money': float(_v[0]),
                    'sy_money': float(_v[1]),
                    'fs_money': float(_v[2]),
                })
            if index == 12:
                _v = dd.split('/')
                if len(_v) != 1:
                    statuerror = True
                    break
                _data.update({
                    'sfqtfy_money': float(_v[0]),
                })
            if index == 13:
                _v = dd.split('/')
                if len(_v) != 3:
                    statuerror = True
                    break
                _data.update({
                    'czyh_money': float(_v[0]),
                    'hd_money': float(_v[1]),
                    'hb_money': float(_v[2]),
                })
            if index == 14:
                _v = dd.split('/')
                if len(_v) != 1:
                    statuerror = True
                    break
                _data.update({
                    'yk_mongey': float(_v[0]),
                })

        if statuerror:
            return self.xtjson.json_params_error('数据解析错误！')
        now = _data.get('form_date')
        start_tiem = datetime.datetime(now.year, now.month, now.day, 0, 0, 0)
        end_time = datetime.datetime(now.year, now.month, now.day, 23, 59, 59)
        if self.MCLS.find_one({'$gte': start_tiem, '$lte': end_time}):
            return self.xtjson.json_result()

        self.MCLS.insert_one(_data)
        return self.xtjson.json_result()

    def export_data_func(self, apth):
        if not os.path.exists(apth):
            return self.xtjson.json_params_error('文件不存在！')

        wb = load_workbook(apth)
        wa = wb.active

        if wa.max_row < 2 or wa.max_column != 15:
            return self.xtjson.json_params_error('文件格式错误！')

        _data = {}
        statuerror = False
        for index in range(15):
            vals = wa.cell(row=2, column=index+1).value
            if isinstance(vals, str):
                vals = vals.replace(',','').replace(' ','')
            if index == 0:
                if isinstance(vals, datetime.datetime):
                    _data.update({
                        'form_date': vals,
                    })
                else:
                    try:
                        form_date = datetime.datetime.strptime(vals, '%Y-%m-%d')
                    except:
                        statuerror = True
                        break
                    _data.update({
                        'form_date': form_date,
                    })
            if index == 1:
                _v = vals.split('/')
                if len(_v) != 3:
                    statuerror = True
                    break
                _data.update({
                    'data_date': _v[0],
                    'app_regist': _v[1],
                    'shou_chong': _v[2],
                })
            if index == 2:
                _v = vals.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'xzyxhy_cp': _v[0],
                    'xzyxhy_dsfyx': _v[1],
                })
            if index == 3:
                try:
                    _v = float(vals)
                    _data.update({
                        'yxyh_count': _v,
                    })
                except:
                    statuerror = True
                    break
            if index == 4:
                _v = vals.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'yxyh_cp_count': _v[0],
                    'yxyh_sfyx_count': _v[1],
                })
            if index == 5:
                _v = vals.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'cpye_money': float(_v[0]),
                    'sfye_money': float(_v[1]),
                })
            if index == 6:
                _v = vals.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'recharge_money': float(_v[0]),
                    'withdraw_money': float(_v[1]),
                })
            if index == 7:
                _v = vals.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'touzhu_money': float(_v[0]),
                    'tsze_money': float(_v[1]),
                })
            if index == 8:
                try:
                    _v = float(vals)
                    _data.update({
                        'sjsy_money': _v,
                    })
                except:
                    statuerror = True
                    break
            if index == 9:
                _v = vals.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'agent_odds': float(_v[0]),
                    'tuishui_money': float(_v[1]),
                })
            if index == 10:
                _v = vals.split('/')
                if len(_v) != 2:
                    statuerror = True
                    break
                _data.update({
                    'jackpot': float(_v[0]),
                    'jackBonus': float(_v[1]),
                })
            if index == 11:
                _v = vals.split('/')
                if len(_v) != 3:
                    statuerror = True
                    break
                _data.update({
                    'sfyx_money': float(_v[0]),
                    'sy_money': float(_v[1]),
                    'fs_money': float(_v[2]),
                })
            if index == 12:
                try:
                    _v = float(vals)
                    _data.update({
                        'sfqtfy_money': _v,
                    })
                except:
                    statuerror = True
                    break
            if index == 13:
                _v = vals.split('/')
                if len(_v) != 3:
                    statuerror = True
                    break
                _data.update({
                    'czyh_money': float(_v[0]),
                    'hd_money': float(_v[1]),
                    'hb_money': float(_v[2]),
                })
            if index == 14:
                try:
                    _v = float(vals)
                    _data.update({
                        'yk_mongey': _v,
                    })
                except:
                    statuerror = True
                    break
        if statuerror:
            return self.xtjson.json_params_error('数据解析错误！')

        self.MCLS.insert_one(_data)
        return self.xtjson.json_result(message='数据导入成功！')

    def post_other_way(self):
        if self.action == 'add_form_data':
            return self.add_form_data_func()
        if self.action == 'importData':
            fileobj = request.files.get('upload')
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('xlsx'):
                return self.xtjson.json_params_error('文件格式错误，只支持xlsx文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/importFile/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)

            return self.export_data_func(import_folder + new_filename + fext)
        if self.action == 'del_all':
            self.MCLS.delete_many({})
            return self.xtjson.json_result()

    def post_data_other_way(self):
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()



class RechargeView(DataFormViewBase):
    title = '首充报表'
    show_menu = False
    add_url_rules = [['/recharge/', 'recharge_log']]
    template = 'data_form/recharge.html'
    MCLS = RechargeTable
    per_page = 30

    def dataform_other_way(self):
        if self.action == 'uploadForm':
            fugai = self.request_data.get('fugai')
            formPath = self.request_data.get('formPath')
            form_date_low = self.request_data.get('form_date')
            if not formPath or not form_date_low:
                return self.xtjson.json_params_error()
            import_folder = current_app.static_folder + '/' + self.project_name + formPath
            if not os.path.exists(import_folder):
                return self.xtjson.json_params_error('文件不存在！')
            try:
                form_date = datetime.datetime.strptime(form_date_low, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('日期错误！')
            try:
                dataf = open(import_folder, 'r', encoding='utf8').read()
            except:
                try:
                    dataf = open(import_folder, 'r', encoding='gbk').read()
                except:
                    return self.xtjson.json_params_error('读取文件失败！')

            _datas = dataf.split('"""\n')
            one_row_data = _datas[0]
            if '处理日期' not in one_row_data or '订单日期' not in one_row_data or '上级代理' not in one_row_data or '付款方式' not in one_row_data:
                return self.xtjson.json_params_error('内容错误！')

            start_time = form_date
            end_time = datetime.datetime(form_date.year, form_date.month, form_date.day, 23, 59, 59)
            datas = []
            for dd in _datas[1:]:
                _v = dd.replace('"','').replace('=','').strip()
                if not _v:
                    continue
                _vv = _v.split(',')
                if not _vv or not _vv[0]:
                    continue
                _d = {
                    'bill_time': datetime.datetime.strptime(_vv[0], '%Y-%m-%d %H:%M:%S'),
                    'cl_time': datetime.datetime.strptime(_vv[1], '%Y-%m-%d %H:%M:%S'),
                    'account': _vv[2] or '',
                    'parent_agent': _vv[3] or '',
                    'username': _vv[4] or '',
                    'over_money': float(_vv[5]) or 0,
                    'order_number': _vv[6] or '',
                    'pay_way': _vv[7] or '',
                    'jy_money': float(_vv[8]) or 0,
                    'operator_user': _vv[9] or '',
                    'collect_money_info': _vv[10] or '',
                    'deposit_money_info': _vv[11] or '',
                    'state': _vv[12] or '',
                    'cz_level': _vv[13] or '',
                    'note': _vv[14] or '',
                    'scdd': _vv[15] or '',
                    'form_date': form_date
                }
                if self.MCLS.find_one({'account': _d.get('account'), 'cl_time': _d.get('cl_time'), 'form_date': {'$gte': start_time, '$lte': end_time}}):
                    if not fugai or fugai == 'false':
                        _result = {
                            'form_date': form_date_low,
                            'formPath': formPath,
                            'cfState': True,
                        }
                        return self.xtjson.json_result(data=_result)
                datas.append(_d)
            for dd in datas:
                _low_d = self.MCLS.find_one({'account': dd.get('account'), 'cl_time': dd.get('cl_time'), 'form_date': {'$gte': start_time, '$lte': end_time}}) or {}
                if _low_d:
                    _low_d.update(dd)
                    self.MCLS.save(_low_d)
                else:
                    self.MCLS.insert_one(dd)
            return self.xtjson.json_result()



class DataFormAnalysisView(CmsFormViewBase):
    title = '分析报表'
    show_menu = False
    add_url_rules = [['/dataForm/analysis', 'dataForm_analysis']]
    template = 'data_form/dataform_fenxi.html'

    def format_money(self, data):
        try:
            return format(data, ",")
        except:
            return data

    def format_datetime(self, data, strft=''):
        if strft:
            try:
                return data.strftime(strft)
            except:
                return data
        if isinstance(data, datetime.datetime):
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    # 代理报表操作
    def dl_table_func(self, page=1, per_page=20, filter={}):
        if not filter.get('form_date'):
            now = datetime.datetime.now() - datetime.timedelta(days=1)
            start_tiem = datetime.datetime(now.year, now.month, now.day, 0, 0, 0)
            end_time = datetime.datetime(now.year, now.month, now.day, 23, 59, 59)
            filter['form_date'] = {'$gte': start_tiem, '$lte': end_time}

        rech_datas = RechargeTable.find_many({'form_date': filter.get('form_date')})
        crech_datas = [ i.get('parent_agent') for i in rech_datas ]

        hyzj_datas = hy_zijing_form_table.find_many({'form_date': filter.get('form_date'), 'chongzhi_money': {'$gt': 0}})
        c_hyzj_datas = [ i.get('parent_agent') for i in hyzj_datas ]

        skip = (page - 1) * per_page
        cp_datas = agency_caipiao_form_table.find_many(filter, sort=[['form_date', -1]])
        sfyx_datas = agency_sanfang_form_table.find_many(filter, sort=[['form_date', -1]])
        tz_datas = agency_touzhu_form_table.find_many(filter, sort=[['form_date', -1]])
        zj_datas = agency_zijing_form_table.find_many(filter, sort=[['form_date', -1]])

        # 合并投注人数
        ddsls = agency_touzhu_form_table.collection().aggregate([
            {"$match": filter},
            {"$group": {"_id": "$account", "count": {"$sum": '$customer_count'}}},
        ])
        hbtz_ls = {}
        for df in ddsls:
            hbtz_ls[df.get('_id')] = df.get('count') or 0

        account_ls = []
        for td in cp_datas:
            account_ls.append(td.get('account'))
        for td in sfyx_datas:
            account_ls.append(td.get('account'))
        for td in tz_datas:
            account_ls.append(td.get('account'))
        for td in zj_datas:
            account_ls.append(td.get('account'))
        account_ls = list(set(account_ls))
        total = len(account_ls)

        pages, total_page = PagingCLS.ustom_pagination(page, total, per_page)
        _account_ls = account_ls[skip: skip+per_page]

        all_datas = []
        for account in _account_ls:
            _data = {'account': account}

            xzzc_datas = CustomerTable.count({'new_time': filter.get('form_date'), 'agency': account}) or 0
            _data['xzzc_count'] = xzzc_datas

            _data['sc_count'] = crech_datas.count(account) or 0
            _data['czry_count'] = c_hyzj_datas.count(account) or 0
            _data['hbtz_count'] = hbtz_ls.get(account) or 0

            # 新增人数和流失人数
            cus_datas1 = hy_zijing_form_table.find_many({'form_date': filter.get('form_date'), 'parent_agent': account})
            dl_cu_ls1 = [i.get('account') for i in cus_datas1]
            start_date = filter.get('form_date').get('$gte')
            end_date = filter.get('form_date').get('$lte')
            dddy = (start_date - end_date).days

            start_db = start_date + datetime.timedelta(days=dddy)
            end_db = end_date + datetime.timedelta(days=dddy)
            cus_datas2 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_db, '$lte': end_db}, 'parent_agent': account})
            dl_cu_ls2 = [i.get('account') for i in cus_datas2]
            new_count = 0
            low_count = 0
            for dl in dl_cu_ls1:
                if dl not in dl_cu_ls2:
                    new_count += 1
            for dl in dl_cu_ls2:
                if dl not in dl_cu_ls1:
                    low_count += 1

            _data['new_count'] = new_count
            _data['low_count'] = low_count

            _d1 = agency_caipiao_form_table.find_many({'form_date': filter.get('form_date'), 'account': account}) or []
            _d2 = agency_sanfang_form_table.find_many({'form_date': filter.get('form_date'), 'account': account}) or []
            _d3 = agency_zijing_form_table.find_many({'form_date': filter.get('form_date'), 'account': account}) or []
            for d1 in _d1:
                xzhy_count = _data.get('xzhy_count') or 0
                xzhy_count += d1.get('xzhy_count')
                betting_count = _data.get('betting_count') or 0
                betting_count += d1.get('betting_count')
                betting_money = _data.get('betting_money') or 0
                betting_money += d1.get('betting_money')
                actual_winlose_money = _data.get('actual_winlose_money') or 0
                actual_winlose_money += d1.get('actual_winlose_money')
                _data.update({
                    'xzhy_count': xzhy_count,
                    'betting_count': betting_count,
                    'betting_money': betting_money,
                    'actual_winlose_money': actual_winlose_money,
                })
            for d2 in _d2:
                bs_number = _data.get('bs_number') or 0
                bs_number += d2.get('bs_number')
                hy_count = _data.get('hy_count') or 0
                hy_count += d2.get('hy_count')
                yx_betting_money = _data.get('yx_betting_money') or 0
                yx_betting_money += d2.get('yx_betting_money')
                hyfs_money = _data.get('hyfs_money') or 0
                hyfs_money += d2.get('hyfs_money')
                sjsy_money = _data.get('sjsy_money') or 0
                sjsy_money += d2.get('sjsy_money')

                _data.update({
                    'bs_number': bs_number,
                    'hy_count': hy_count,
                    'yx_betting_money': yx_betting_money,
                    'hyfs_money': hyfs_money,
                    'sjsy_money': sjsy_money,
                })
            for d3 in _d3:
                chongzhi_money = _data.get('chongzhi_money') or 0
                chongzhi_money += d3.get('chongzhi_money') or 0
                hdjq_money = _data.get('hdjq_money') or 0
                hdjq_money += d3.get('hdjq_money') or 0
                tixian_money = _data.get('tixian_money') or 0
                tixian_money += d3.get('tixian_money') or 0
                hdkq_money = _data.get('hdkq_money') or 0
                hdkq_money += d3.get('hdkq_money') or 0
                hd_money = _data.get('hd_money') or 0
                hd_money += d3.get('hd_money') or 0
                hb_money = _data.get('hb_money') or 0
                hb_money += d3.get('hb_money') or 0
                czyh_money = _data.get('czyh_money') or 0
                czyh_money += d3.get('czyh_money') or 0

                _data.update({
                    'chongzhi_money': chongzhi_money,
                    'hdjq_money': hdjq_money,
                    'tixian_money': tixian_money,
                    'hdkq_money': hdkq_money,
                    'hd_money': hd_money,
                    'hb_money': hb_money,
                    'czyh_money': czyh_money,
                })

            all_datas.append(_data)

        table_html = f'''
                                <tr>
                                    <td>账号[{ total }]</td>
                                    <td>彩票下注会员</td>
                                    <td>彩票笔数</td>
                                    <td>投注金额</td>
                                    <td>彩票输赢</td>
                                    <td>三方笔数</td>
                                    <td>三方下注会员数</td>
                                    <td>三方有效投注金额</td>
                                    <td>返水金额</td>
                                    <td>三方实际输赢</td>
                                    <td>充值金额</td>
                                    <td>后台加钱</td>
                                    <td>提现金额</td>
                                    <td>后台扣钱</td>
                                    <td>活动金额</td>
                                    <td>红包金额</td>
                                    <td>手续费</td>
                                    <td>新增人数</td>
                                    <td>流失人数</td>
                                    <td>首充人数</td>
                                    <td>充值人数</td>
                                    <td>新增注册</td>
                                    <td>合并投注人数</td>
                                </tr>        
        '''
        for data in all_datas:
            table_html += f'''
                                <tr>
                                    <td>{ data.get('account') or '' }</td>
                                    <td>{ data.get('xzhy_count') or 0 }</td>
                                    <td>{ data.get('betting_count') or 0 }</td>
                                    <td>{ data.get('betting_money') or 0 }</td>
                                    <td>{ data.get('actual_winlose_money') or 0 }</td>
                                    <td>{ data.get('bs_number') or 0 }</td>
                                    <td>{ data.get('hy_count') or 0 }</td>
                                    <td>{ data.get('yx_betting_money') or 0 }</td>
                                    <td>{ data.get('hyfs_money') or 0 }</td>
                                    <td>{ data.get('sjsy_money') or 0 }</td>
                                    <td>{ data.get('chongzhi_money') or 0 }</td>
                                    <td>{ data.get('hdjq_money') or 0 }</td>
                                    <td>{ data.get('tixian_money') or 0 }</td>
                                    <td>{ data.get('hdkq_money') or 0 }</td>
                                    <td>{ data.get('hd_money') or 0 }</td>
                                    <td>{ data.get('hb_money') or 0 }</td>
                                    <td>{ data.get('czyh_money') or 0 }</td>
                                    <td>{ data.get('new_count') or 0 }</td>
                                    <td>{ data.get('low_count') or 0 }</td>
                                    <td>{ data.get('sc_count') or 0 }</td>
                                    <td>{ data.get('czry_count') or 0 }</td>
                                    <td>{ data.get('xzzc_count') or 0 }</td>
                                    <td>{ data.get('hbtz_count') or 0 }</td>
                                </tr>                 
            '''

        page_nav_html = ''
        if total == 0:
            page_nav_html += '''
                <li class="page-item disabled">
                    <a class="page-link" aria-label="Previous">
                        <span aria-hidden="true">&laquo;</span>
                    </a>
                </li>   
                <li class="page-item disabled">
                    <a class="page-link" aria-label="Next">
                        <span aria-hidden="true">&raquo;</span>
                    </a>
                </li>
            '''
        else:
            if page == 1:
                page_nav_html += f'''
                    <li class="page-item disabled">
                        <a class="page-link" aria-label="Previous">
                            <span aria-hidden="true">&laquo;</span>
                        </a>
                    </li>
                '''
            else:
                page_nav_html += f'''
                    <li class="page-item" onclick="dl_request_data({page - 1})">
                        <a class="page-link" aria-label="Previous">
                            <span aria-hidden="true">&laquo;</span>
                        </a>
                    </li>           
                '''

            for crrpage in pages:
                if crrpage == page:
                    page_nav_html += f'''
                    <li class="page-item active"><a class="page-link">{crrpage}</a></li>
                    '''
                else:
                    page_nav_html += f'''
                    <li class="page-item" onclick="dl_request_data({crrpage})"><a class="page-link">{crrpage}</a></li>
                    '''

            if page == total_page:
                page_nav_html += f'''
                    <li class="page-item disabled">
                        <a class="page-link" aria-label="Next">
                            <span aria-hidden="true">&raquo;</span>
                        </a>
                    </li>                    
                '''
            else:
                page_nav_html += f'''
                    <li class="page-item" onclick="dl_request_data({page + 1})">
                        <a class="page-link" aria-label="Next">
                            <span aria-hidden="true">&raquo;</span>
                        </a>
                    </li>               
                '''

        dataTableBottom_html = f'''   
        <span>总信息条数：{total}，总{total_page}页</span>
        <ul class="pages pagination" data-crrpage="{page}">
            {page_nav_html}                     
        </ul>    
        '''

        return all_datas, total, pages, total_page, table_html, dataTableBottom_html

    def get_dl_account_sum_data(self, account, start_date, end_date):
        _data = {'account': account}
        _d1 = agency_caipiao_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []
        _d2 = agency_sanfang_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []
        _d3 = agency_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []

        rech_datas = RechargeTable.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'parent_agent': account}) or []
        hyzj_datas = hy_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'chongzhi_money': {'$gt': 0}, 'parent_agent': account}) or []
        _data['sc_count'] = len(rech_datas)
        _data['czry_count'] = len(hyzj_datas)

        xzzc_datas = CustomerTable.count({'new_time': {'$gte': start_date, '$lte': end_date}, 'agency': account}) or 0
        _data['xzzc_count'] = xzzc_datas

        cus_datas1 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'parent_agent': account})
        dl_cu_ls1 = [i.get('account') for i in cus_datas1]
        dddy = (start_date - end_date).days

        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        cus_datas2 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_db, '$lte': end_db}, 'parent_agent': account})
        dl_cu_ls2 = [i.get('account') for i in cus_datas2]
        new_count = 0
        low_count = 0
        for dl in dl_cu_ls1:
            if dl not in dl_cu_ls2:
                new_count += 1
        for dl in dl_cu_ls2:
            if dl not in dl_cu_ls1:
                low_count += 1
        _data['new_count'] = new_count
        _data['low_count'] = low_count

        # 合并投注人数
        ddsls = agency_touzhu_form_table.collection().aggregate([
            {"$match": {'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}},
            {"$group": {"_id": "$account", "count": {"$sum": '$customer_count'}}},
        ])
        hbtz_ls = {}
        for df in ddsls:
            hbtz_ls[df.get('_id')] = df.get('count') or 0

        _data['hbtz_count'] = hbtz_ls.get(account) or 0

        for d1 in _d1:
            xzhy_count = _data.get('xzhy_count') or 0
            xzhy_count += d1.get('xzhy_count')
            betting_count = _data.get('betting_count') or 0
            betting_count += d1.get('betting_count')
            betting_money = _data.get('betting_money') or 0
            betting_money += d1.get('betting_money')
            actual_winlose_money = _data.get('actual_winlose_money') or 0
            actual_winlose_money += d1.get('actual_winlose_money')
            _data.update({
                'xzhy_count': xzhy_count,
                'betting_count': betting_count,
                'betting_money': betting_money,
                'actual_winlose_money': actual_winlose_money,
            })
        for d2 in _d2:
            bs_number = _data.get('bs_number') or 0
            bs_number += d2.get('bs_number')
            hy_count = _data.get('hy_count') or 0
            hy_count += d2.get('hy_count')
            yx_betting_money = _data.get('yx_betting_money') or 0
            yx_betting_money += d2.get('yx_betting_money')
            hyfs_money = _data.get('hyfs_money') or 0
            hyfs_money += d2.get('hyfs_money')
            sjsy_money = _data.get('sjsy_money') or 0
            sjsy_money += d2.get('sjsy_money')

            _data.update({
                'bs_number': bs_number,
                'hy_count': hy_count,
                'yx_betting_money': yx_betting_money,
                'hyfs_money': hyfs_money,
                'sjsy_money': sjsy_money,
            })
        for d3 in _d3:
            chongzhi_money = _data.get('chongzhi_money') or 0
            chongzhi_money += d3.get('chongzhi_money') or 0
            hdjq_money = _data.get('hdjq_money') or 0
            hdjq_money += d3.get('hdjq_money') or 0
            tixian_money = _data.get('tixian_money') or 0
            tixian_money += d3.get('tixian_money') or 0
            hdkq_money = _data.get('hdkq_money') or 0
            hdkq_money += d3.get('hdkq_money') or 0
            hd_money = _data.get('hd_money') or 0
            hd_money += d3.get('hd_money') or 0
            hb_money = _data.get('hb_money') or 0
            hb_money += d3.get('hb_money') or 0
            czyh_money = _data.get('czyh_money') or 0
            czyh_money += d3.get('czyh_money') or 0

            _data.update({
                'chongzhi_money': chongzhi_money,
                'hdjq_money': hdjq_money,
                'tixian_money': tixian_money,
                'hdkq_money': hdkq_money,
                'hd_money': hd_money,
                'hb_money': hb_money,
                'czyh_money': czyh_money,
            })
        return _data

    def dbfx_func(self, account, start_date, end_date):
        _data1 = self.get_dl_account_sum_data(account, start_date, end_date)
        dddy = (start_date - end_date).days
        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        _data2 = self.get_dl_account_sum_data(account, start_db, end_db)

        fields = [
            'xzhy_count',
            'betting_count',
            'betting_money',
            'actual_winlose_money',
            'bs_number',
            'hy_count',
            'yx_betting_money',
            'hyfs_money',
            'sjsy_money',
            'chongzhi_money',
            'hdjq_money',
            'tixian_money',
            'hdkq_money',
            'hd_money',
            'hb_money',
            'czyh_money',
            'sc_count',
            'czry_count',
            'xzzc_count',
            'hbtz_count',
            'new_count',
            'low_count',
        ]
        datas = []
        for filed in fields:
            name = ''
            if filed == 'xzhy_count':
                name = '彩票下注会员'
            if filed == 'betting_count':
                name = '彩票笔数'
            if filed == 'betting_money':
                name = '投注金额'
            if filed == 'actual_winlose_money':
                name = '彩票输赢'
            if filed == 'bs_number':
                name = '三方笔数'
            if filed == 'hy_count':
                name = '三方下注会员数'
            if filed == 'yx_betting_money':
                name = '三方有效投注金额'
            if filed == 'hyfs_money':
                name = '返水金额'
            if filed == 'sjsy_money':
                name = '三方实际输赢'
            if filed == 'chongzhi_money':
                name = '充值金额'
            if filed == 'hdjq_money':
                name = '后台加钱'
            if filed == 'tixian_money':
                name = '提现金额'
            if filed == 'hdkq_money':
                name = '后台扣钱'
            if filed == 'hd_money':
                name = '活动金额'
            if filed == 'hb_money':
                name = '红包金额'
            if filed == 'czyh_money':
                name = '手续费'
            if filed == 'sc_count':
                name = '首充人数'
            if filed == 'czry_count':
                name = '充值人数'
            if filed == 'xzzc_count':
                name = '新增注册'
            if filed == 'hbtz_count':
                name = '合并投注人数'
            if filed == 'new_count':
                name = '新增人数'
            if filed == 'low_count':
                name = '流失人数'

            v1 = _data1.get(filed) or 0
            v2 = _data2.get(filed) or 0
            if v2 == 0 and v1 == 0:
                vv = 0
            elif v2 == 0:
                vv = 100
            else:
                vv = round((v1 - v2) / v2 * 100, 2)
            datas.append({
                'name': name,
                'v1': v1,
                'v2': v2,
                'vv': vv,
            })

        td_html = ''
        for index, d in enumerate(datas):
            last_item = False
            if index == (len(datas) - 1):
                last_item = True

            click_html = 'onclick="export_ls_func({\'action\': \'export_ls\', \'start_date\':\'%s\', \'end_date\':\'%s\', \'account\': \'%s\'})"' % (
                start_date.strftime('%Y-%m-%d %H:%M:%S'),
                end_date.strftime('%Y-%m-%d %H:%M:%S'),
                account
            )
            if d.get('name') in ['新增人数', '流失人数']:
                td_html += f'''
                        <tr { click_html if last_item else '' }>
                            <td>{ d.get('name') or '' }</td>
                            <td>{ d.get('v1') or 0 }</td>
                            <td></td>
                            <td></td>
                        </tr>                     
                '''
            else:
                td_html += f'''
                        <tr { click_html if last_item else '' }>
                            <td>{ d.get('name') or '' }</td>
                            <td>{ d.get('v1') or 0 }</td>
                            <td>{ d.get('v2') or 0 }</td>
                            <td>{ d.get('vv') }%</td>
                        </tr>                     
                '''

        html = f'''
            <table class="table table-bordered table-hover text-center" style="background-color: #ffffff; font-size: 13px;">
                <tbody>
                    <tr>
                        <td>名称</td>
                        <td>{'%s 至 %s' % (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))}</td>
                        <td>{'%s 至 %s' % (start_db.strftime('%Y-%m-%d'), end_db.strftime('%Y-%m-%d'))}</td>
                        <td>增加百分比</td>
                    </tr>
                    { td_html }
                </tbody>
            </table>        
        '''
        return self.xtjson.json_result(message=html, data={'title': '%s, 对比分析结果' % account})

    def get_db_html(self, _w_actionm, history_date=''):
        if _w_actionm == 'yx_dbfx':
            html = f"""
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>游戏：</span></div>
                    <input type="text" class="form-control" id="game_name" placeholder="游戏">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">上传日期：</span></div>
                    <input type="text" class="form-control pickerdate" value="{history_date or ''}" onmouseenter="$.picker_YY_HH_DD_HH_MM_SS('.pickerdate')" id="db_date" placeholder="上传日期">
                </div>
                <span class="btn btn-primary swal2-styled" onclick="post_db_func('{_w_actionm}')" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>                        
            """
        else:
            html = f"""
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>账户：</span></div>
                    <input type="text" class="form-control" id="db_account" placeholder="账户">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">上传日期：</span></div>
                    <input type="text" class="form-control pickerdate" value="{ history_date or '' }" onmouseenter="$.picker_YY_HH_DD_HH_MM_SS('.pickerdate')" id="db_date" placeholder="上传日期">
                </div>
                <span class="btn btn-primary swal2-styled" onclick="post_db_func('{_w_actionm}')" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>                        
            """
        return self.xtjson.json_result(message=html)

    # 会员报表操作
    def hy_table_func(self, page=1, per_page=20, filter={}):
        if not filter.get('form_date'):
            now = datetime.datetime.now() - datetime.timedelta(days=1)
            start_tiem = datetime.datetime(now.year, now.month, now.day, 0, 0, 0)
            end_time = datetime.datetime(now.year, now.month, now.day, 23, 59, 59)
            filter['form_date'] = {'$gte': start_tiem, '$lte': end_time}

        skip = (page - 1) * per_page
        cp_datas = hy_caipiao_form_tabl.find_many(filter, sort=[['form_date', -1]])
        sfyx_datas = hy_sanfang_form_table.find_many(filter, sort=[['form_date', -1]])
        zj_datas = hy_zijing_form_table.find_many(filter, sort=[['form_date', -1]])

        account_ls = []
        ad_dict = {}
        for td in cp_datas:
            account = td.get('account')
            ad_dict[account] = td.get('parent_agent')
            account_ls.append(account)
        for td in sfyx_datas:
            account = td.get('account')
            ad_dict[account] = td.get('parent_agent')
            account_ls.append(account)
        for td in zj_datas:
            account = td.get('account')
            ad_dict[account] = td.get('parent_agent')
            account_ls.append(account)
        account_ls = list(set(account_ls))
        total = len(account_ls)

        pages, total_page = PagingCLS.ustom_pagination(page, total, per_page)
        _account_ls = account_ls[skip: skip+per_page]

        all_datas = []
        for account in _account_ls:
            _data = {'account': account}
            _data['parent_agent'] = ad_dict.get(account) or ''
            _d1 = hy_caipiao_form_tabl.find_many({'form_date': filter.get('form_date'), 'account': account}) or []
            _d2 = hy_sanfang_form_table.find_many({'form_date': filter.get('form_date'), 'account': account}) or []
            _d3 = hy_zijing_form_table.find_many({'form_date': filter.get('form_date'), 'account': account}) or []
            for d1 in _d1:
                betting_count = _data.get('betting_count') or 0
                betting_count += d1.get('betting_count')
                betting_money = _data.get('betting_money') or 0
                betting_money += d1.get('betting_money')
                actual_winlose_money = _data.get('actual_winlose_money') or 0
                actual_winlose_money += d1.get('actual_winlose_money')

                _data.update({
                    'betting_count': betting_count,
                    'betting_money': betting_money,
                    'actual_winlose_money': actual_winlose_money,
                })
            for d2 in _d2:
                bs_number = _data.get('bs_number') or 0
                bs_number += d2.get('bs_number')
                yx_betting_money = _data.get('yx_betting_money') or 0
                yx_betting_money += d2.get('yx_betting_money')
                fs_money = _data.get('fs_money') or 0
                fs_money += d2.get('fs_money')
                sf_actual_winlose_money = _data.get('sf_actual_winlose_money') or 0
                sf_actual_winlose_money += d2.get('actual_winlose_money')

                _data.update({
                    'bs_number': bs_number,
                    'yx_betting_money': yx_betting_money,
                    'fs_money': fs_money,
                    'sf_actual_winlose_money': sf_actual_winlose_money,
                })
            for d3 in _d3:
                chongzhi_money = _data.get('chongzhi_money') or 0
                chongzhi_money += d3.get('chongzhi_money') or 0
                xscz_money = _data.get('xscz_money') or 0
                xscz_money += d3.get('xscz_money') or 0
                xxcz_money = _data.get('xxcz_money') or 0
                xxcz_money += d3.get('xxcz_money') or 0
                hdjq_money = _data.get('hdjq_money') or 0
                hdjq_money += d3.get('hdjq_money') or 0
                tixian_money = _data.get('tixian_money') or 0
                tixian_money += d3.get('tixian_money') or 0
                hdkq_money = _data.get('hdkq_money') or 0
                hdkq_money += d3.get('hdkq_money') or 0
                hd_money = _data.get('hd_money') or 0
                hd_money += d3.get('hd_money') or 0
                hb_money = _data.get('hb_money') or 0
                hb_money += d3.get('hb_money') or 0
                czyh_money = _data.get('czyh_money') or 0
                czyh_money += d3.get('czyh_money') or 0
                hb_money = _data.get('hb_money') or 0
                hb_money += d3.get('hb_money') or 0

                _data.update({
                    'chongzhi_money': chongzhi_money,
                    'xscz_money': xscz_money,
                    'xxcz_money': xxcz_money,
                    'hdjq_money': hdjq_money,
                    'tixian_money': tixian_money,
                    'hdkq_money': hdkq_money,
                    'hd_money': hd_money,
                    'hb_money': hb_money,
                    'czyh_money': czyh_money,
                })

            all_datas.append(_data)

        table_html = f'''
                                <tr>
                                    <td>账号[{ total }]</td>
                                    <td>上级代理</td>
                                    <td>彩票笔数</td>
                                    <td>投注金额</td>
                                    <td>彩票输赢</td>
                                    <td>三方笔数</td>
                                    <td>三方有效投注金额</td>
                                    <td>返水金额</td>
                                    <td>三方实际输赢</td>
                                    <td>充值金额</td>
                                    <td>线上充值</td>
                                    <td>线下充值</td>
                                    <td>后台加钱</td>
                                    <td>提现金额</td>
                                    <td>后台扣钱</td>
                                    <td>活动金额</td>
                                    <td>红包金额</td>
                                    <td>手续费</td>
                                </tr>        
        '''
        for data in all_datas:
            table_html += f'''
                                <tr>
                                    <td>{ data.get('account') or '' }</td>
                                    <td>{ data.get('parent_agent') or '' }</td>
                                    <td>{ data.get('betting_count') or 0 }</td>
                                    <td>{ data.get('betting_money') or 0 }</td>
                                    <td>{ data.get('actual_winlose_money') or 0 }</td>
                                    <td>{ data.get('bs_number') or 0 }</td>
                                    <td>{ data.get('yx_betting_money') or 0 }</td>
                                    <td>{ data.get('fs_money') or 0 }</td>
                                    <td>{ data.get('sf_actual_winlose_money') or 0 }</td>
                                    <td>{ data.get('chongzhi_money') or 0 }</td>
                                    <td>{ data.get('xscz_money') or 0 }</td>
                                    <td>{ data.get('xxcz_money') or 0 }</td>
                                    <td>{ data.get('hdjq_money') or 0 }</td>
                                    <td>{ data.get('tixian_money') or 0 }</td>
                                    <td>{ data.get('hdkq_money') or 0 }</td>
                                    <td>{ data.get('hd_money') or 0 }</td>
                                    <td>{ data.get('hb_money') or 0 }</td>
                                    <td>{ data.get('czyh_money') or 0 }</td>
                                </tr>                 
            '''

        page_nav_html = ''
        if total == 0:
            page_nav_html += '''
                <li class="page-item disabled">
                    <a class="page-link" aria-label="Previous">
                        <span aria-hidden="true">&laquo;</span>
                    </a>
                </li>   
                <li class="page-item disabled">
                    <a class="page-link" aria-label="Next">
                        <span aria-hidden="true">&raquo;</span>
                    </a>
                </li>
            '''
        else:
            if page == 1:
                page_nav_html += f'''
                    <li class="page-item disabled">
                        <a class="page-link" aria-label="Previous">
                            <span aria-hidden="true">&laquo;</span>
                        </a>
                    </li>
                '''
            else:
                page_nav_html += f'''
                    <li class="page-item" onclick="hy_request_daat({page - 1})">
                        <a class="page-link" aria-label="Previous">
                            <span aria-hidden="true">&laquo;</span>
                        </a>
                    </li>           
                '''

            for crrpage in pages:
                if crrpage == page:
                    page_nav_html += f'''
                    <li class="page-item active"><a class="page-link">{crrpage}</a></li>
                    '''
                else:
                    page_nav_html += f'''
                    <li class="page-item" onclick="hy_request_daat({crrpage})"><a class="page-link">{crrpage}</a></li>
                    '''

            if page == total_page:
                page_nav_html += f'''
                    <li class="page-item disabled">
                        <a class="page-link" aria-label="Next">
                            <span aria-hidden="true">&raquo;</span>
                        </a>
                    </li>                    
                '''
            else:
                page_nav_html += f'''
                    <li class="page-item" onclick="hy_request_daat({page + 1})">
                        <a class="page-link" aria-label="Next">
                            <span aria-hidden="true">&raquo;</span>
                        </a>
                    </li>               
                '''

        dataTableBottom_html = f'''   
        <span>总信息条数：{total}，总{total_page}页</span>
        <ul class="pages pagination" data-crrpage="{page}">
            {page_nav_html}                     
        </ul>    
        '''

        return all_datas, total, pages, total_page, table_html, dataTableBottom_html

    def get_hy_account_sum_data(self, account, start_date, end_date):
        _data = {'account': account}
        _d1 = hy_caipiao_form_tabl.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []
        _d2 = hy_sanfang_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []
        _d3 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []
        for d1 in _d1:
            betting_count = _data.get('betting_count') or 0
            betting_count += d1.get('betting_count')
            betting_money = _data.get('betting_money') or 0
            betting_money += d1.get('betting_money')
            actual_winlose_money = _data.get('actual_winlose_money') or 0
            actual_winlose_money += d1.get('actual_winlose_money')

            _data.update({
                'betting_count': betting_count,
                'betting_money': betting_money,
                'actual_winlose_money': actual_winlose_money,
            })
        for d2 in _d2:
            bs_number = _data.get('bs_number') or 0
            bs_number += d2.get('bs_number')
            yx_betting_money = _data.get('yx_betting_money') or 0
            yx_betting_money += d2.get('yx_betting_money')
            fs_money = _data.get('fs_money') or 0
            fs_money += d2.get('fs_money')
            sf_actual_winlose_money = _data.get('sf_actual_winlose_money') or 0
            sf_actual_winlose_money += d2.get('actual_winlose_money')

            _data.update({
                'bs_number': bs_number,
                'yx_betting_money': yx_betting_money,
                'fs_money': fs_money,
                'sf_actual_winlose_money': sf_actual_winlose_money,
            })
        for d3 in _d3:
            chongzhi_money = _data.get('chongzhi_money') or 0
            chongzhi_money += d3.get('chongzhi_money') or 0
            xscz_money = _data.get('xscz_money') or 0
            xscz_money += d3.get('xscz_money') or 0
            xxcz_money = _data.get('xxcz_money') or 0
            xxcz_money += d3.get('xxcz_money') or 0
            hdjq_money = _data.get('hdjq_money') or 0
            hdjq_money += d3.get('hdjq_money') or 0
            tixian_money = _data.get('tixian_money') or 0
            tixian_money += d3.get('tixian_money') or 0
            hdkq_money = _data.get('hdkq_money') or 0
            hdkq_money += d3.get('hdkq_money') or 0
            hd_money = _data.get('hd_money') or 0
            hd_money += d3.get('hd_money') or 0
            hb_money = _data.get('hb_money') or 0
            hb_money += d3.get('hb_money') or 0
            czyh_money = _data.get('czyh_money') or 0
            czyh_money += d3.get('czyh_money') or 0
            hb_money = _data.get('hb_money') or 0
            hb_money += d3.get('hb_money') or 0

            _data.update({
                'chongzhi_money': chongzhi_money,
                'xscz_money': xscz_money,
                'xxcz_money': xxcz_money,
                'hdjq_money': hdjq_money,
                'tixian_money': tixian_money,
                'hdkq_money': hdkq_money,
                'hd_money': hd_money,
                'hb_money': hb_money,
                'czyh_money': czyh_money,
            })

        return _data

    def hy_dbfx_func(self, account, start_date, end_date):
        _data1 = self.get_hy_account_sum_data(account, start_date, end_date)
        dddy = (start_date - end_date).days
        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        _data2 = self.get_hy_account_sum_data(account, start_db, end_db)

        fields = [
            'betting_count',
            'betting_money',
            'actual_winlose_money',
            'bs_number',
            'yx_betting_money',
            'fs_money',
            'sf_actual_winlose_money',
            'chongzhi_money',
            'xscz_money',
            'xxcz_money',
            'hdjq_money',
            'tixian_money',
            'hdkq_money',
            'hd_money',
            'hb_money',
            'czyh_money',
        ]
        datas = []
        for filed in fields:
            name = ''
            if filed == 'betting_count':
                name = '彩票笔数'
            if filed == 'betting_money':
                name = '投注金额'
            if filed == 'actual_winlose_money':
                name = '彩票输赢'
            if filed == 'bs_number':
                name = '三方笔数'
            if filed == 'yx_betting_money':
                name = '三方有效投注金额'
            if filed == 'fs_money':
                name = '返水金额'
            if filed == 'sf_actual_winlose_money':
                name = '三方实际输赢'
            if filed == 'chongzhi_money':
                name = '充值金额'
            if filed == 'xscz_money':
                name = '线上充值'
            if filed == 'xxcz_money':
                name = '线下充值'
            if filed == 'hdjq_money':
                name = '后台加钱'
            if filed == 'tixian_money':
                name = '提现金额'
            if filed == 'hdkq_money':
                name = '后台扣钱'
            if filed == 'hd_money':
                name = '活动金额'
            if filed == 'hb_money':
                name = '红包金额'
            if filed == 'czyh_money':
                name = '手续费'

            v1 = _data1.get(filed) or 0
            v2 = _data2.get(filed) or 0
            if v2 == 0 and v1 == 0:
                vv = 0
            elif v2 == 0:
                vv = 100
            else:
                vv = round((v1 - v2) / v2 * 100, 2)
            datas.append({
                'name': name,
                'v1': v1,
                'v2': v2,
                'vv': vv,
            })

        td_html = ''
        for d in datas:
            td_html += f'''
                    <tr>
                        <td>{ d.get('name') or '' }</td>
                        <td>{ d.get('v1') or 0 }</td>
                        <td>{ d.get('v2') or 0 }</td>
                        <td>{ d.get('vv') }%</td>
                    </tr>                     
            '''

        html = f'''
            <table class="table table-bordered table-hover text-center" style="background-color: #ffffff; font-size: 13px;">
                <tbody>
                    <tr>
                        <td>名称</td>
                        <td>{'%s 至 %s' % (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))}</td>
                        <td>{'%s 至 %s' % (start_db.strftime('%Y-%m-%d'), end_db.strftime('%Y-%m-%d'))}</td>
                        <td>增加百分比</td>
                    </tr>
                    { td_html }
                </tbody>
            </table>        
        '''
        return self.xtjson.json_result(message=html, data={'title': '%s, 对比分析结果' % account})

    def hy_jdtz_export(self):
        limit_number = self.request_data.get('limit_number')
        db_date = self.request_data.get('db_date')
        if not limit_number or not db_date:
            return self.xtjson.json_params_error()
        start_date, end_date = PagingCLS.by_silce(db_date)

        hy_caipiao_form_tabl.find_many({'form_date': {'$gte': start_date, '$lte': end_date}})

    def get_hy_tzfx_html(self):
        html = '''
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>限制最低金额：</span></div>
                    <input type="text" class="form-control" id="limit_number" placeholder="限制最低金额">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">时间：</span></div>
                    <input type="text" class="form-control pickerdate" value="" onmouseenter="$.picker_YY_HH_DD_HH_MM_SS('.pickerdate')" id="db_date" placeholder="时间">
                </div>
                <span class="btn btn-primary swal2-styled" onclick="hy_jdtz_export_func()" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>                      
        '''
        return self.xtjson.json_result(data=html)

    def thr_export_ls_func(self, root_path, account, start_date, end_date):
        # 新增人数和流失人数
        cus_datas1 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'parent_agent': account})
        dl_cu_ls1 = [i.get('account') for i in cus_datas1]
        dddy = (start_date - end_date).days

        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        cus_datas2 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_db, '$lte': end_db}, 'parent_agent': account})
        low_accounts = []
        for dl in cus_datas2:
            account = dl.get('account') or ''
            if account not in dl_cu_ls1:
                low_accounts.append(account)

        headers = [
            '用户姓名', '用户账户', '代理', '号码'
        ]
        filead_datas = [
            {
                'db_filed': 'account',
                'type': 'str',
            },{
                'db_filed': 'agency',
                'type': 'str',
            },{
                'db_filed': 'user_tele',
                'type': 'str',
            },{
                'db_filed': 'user_name',
                'type': 'str',
            },
        ]

        absolute_folter = os.path.join(root_path, self.project_static_folder)
        export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
        filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
        export_data = {
            'filename': filename,
            'statu': ExportStatu.ongoing,
            'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
            'total': len(low_accounts),
            'out_count': 0,
            'note': '流失人数导出数据-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        }
        uuid = ExportDataModel.insert_one(export_data)

        datas = []
        for ac in low_accounts:
            _d = {
                'account': ac
            }
            account_data = CustomerTable.find_one({'account': ac}) or {}
            if account_data:
                _d.update({
                    'agency': account_data.get('agency'),
                    'user_tele': account_data.get('user_tele'),
                    'user_name': account_data.get('user_name'),
                })
            datas.append(_d)

        self.export_xlsx_func(uuid, export_folder, filename, datas, headers, filead_datas=filead_datas)

    # 导出流失人数数据
    def export_ls_func(self):
        account = self.request_data.get('account')
        start_date = self.request_data.get('start_date')
        end_date = self.request_data.get('end_date')
        if not account or not start_date or not end_date:
            return self.xtjson.json_params_error()

        start_date, end_date = datetime.datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S'), datetime.datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S')
        threading.Thread(target=self.thr_export_ls_func, args=(current_app.root_path, account, start_date, end_date)).start()
        return self.xtjson.json_result()

    # 游戏报表
    def yx_table_func(self, page=1, per_page=20, filter={}):
        if not filter.get('form_date'):
            now = datetime.datetime.now() - datetime.timedelta(days=1)
            start_tiem = datetime.datetime(now.year, now.month, now.day, 0, 0, 0)
            end_time = datetime.datetime(now.year, now.month, now.day, 23, 59, 59)
            filter['form_date'] = {'$gte': start_tiem, '$lte': end_time}

        df_f_fict = {}
        df_f_fict.update(filter)
        if df_f_fict.get('game_name'):
            df_f_fict['platform'] = df_f_fict.get('game_name')
            df_f_fict.pop('game_name')

        skip = (page - 1) * per_page
        sf_datas = game_sfyx_form_table.find_many(df_f_fict)
        tj_datas = game_yxtj_form_table.find_many(filter)

        game_name_ls = []
        for td in sf_datas:
            game_name = td.get('platform')
            game_name_ls.append(game_name)
        for td in tj_datas:
            game_name = td.get('game_name')
            game_name_ls.append(game_name)
        game_name_ls = list(set(game_name_ls))
        total = len(game_name_ls)

        pages, total_page = PagingCLS.ustom_pagination(page, total, per_page)
        _account_ls = game_name_ls[skip: skip+per_page]

        all_datas = []
        for game_name in _account_ls:
            _data = {'game_name': game_name}
            _f = {}
            df_f_fict['platform'] = game_name
            _d1 = game_sfyx_form_table.find_many(df_f_fict) or []
            filter['game_name'] = game_name
            _d2 = game_yxtj_form_table.find_many(filter) or []
            for d1 in _d1:
                yx_betting_money = _data.get('yx_betting_money') or 0
                yx_betting_money += d1.get('yx_betting_money')
                fs_money = _data.get('fs_money') or 0
                fs_money += d1.get('fs_money')
                sjsy_money = _data.get('sjsy_money') or 0
                sjsy_money += d1.get('sjsy_money')

                _data.update({
                    'yx_betting_money': yx_betting_money,
                    'fs_money': fs_money,
                    'sjsy_money': sjsy_money,
                })
            for d2 in _d2:
                bs_number = _data.get('bs_number') or 0
                bs_number += d2.get('bs_number')
                betting_money = _data.get('betting_money') or 0
                betting_money += d2.get('betting_money')
                actual_winlose_money = _data.get('actual_winlose_money') or 0
                actual_winlose_money += d2.get('actual_winlose_money')
                xzhy_count = _data.get('xzhy_count') or 0
                xzhy_count += d2.get('xzhy_count')

                _data.update({
                    'bs_number': bs_number,
                    'betting_money': betting_money,
                    'actual_winlose_money': actual_winlose_money,
                    'xzhy_count': xzhy_count,
                })
            all_datas.append(_data)

        table_html = f'''
                                <tr>
                                    <td>游戏/平台[{ total }]</td>
                                    <td>有效投注金额</td>
                                    <td>返水金额</td>
                                    <td>实际输赢(含返水)</td>
                                    <td>笔数</td>
                                    <td>投注金额</td>
                                    <td>实际输赢（含退水）</td>
                                    <td>下注会员数</td>
                                </tr>        
        '''
        for data in all_datas:
            table_html += f'''
                                <tr>
                                    <td>{ data.get('game_name') or '' }</td>
                                    <td>{ data.get('yx_betting_money') or 0 }</td>
                                    <td>{ data.get('fs_money') or 0 }</td>
                                    <td>{ data.get('sjsy_money') or 0 }</td>
                                    <td>{ data.get('bs_number') or 0 }</td>
                                    <td>{ data.get('betting_money') or 0 }</td>
                                    <td>{ data.get('actual_winlose_money') or 0 }</td>
                                    <td>{ data.get('xzhy_count') or 0 }</td>
                                </tr>                 
            '''

        page_nav_html = ''
        if total == 0:
            page_nav_html += '''
                <li class="page-item disabled">
                    <a class="page-link" aria-label="Previous">
                        <span aria-hidden="true">&laquo;</span>
                    </a>
                </li>   
                <li class="page-item disabled">
                    <a class="page-link" aria-label="Next">
                        <span aria-hidden="true">&raquo;</span>
                    </a>
                </li>
            '''
        else:
            if page == 1:
                page_nav_html += f'''
                    <li class="page-item disabled">
                        <a class="page-link" aria-label="Previous">
                            <span aria-hidden="true">&laquo;</span>
                        </a>
                    </li>
                '''
            else:
                page_nav_html += f'''
                    <li class="page-item" onclick="yx_request_data({page - 1})">
                        <a class="page-link" aria-label="Previous">
                            <span aria-hidden="true">&laquo;</span>
                        </a>
                    </li>           
                '''

            for crrpage in pages:
                if crrpage == page:
                    page_nav_html += f'''
                    <li class="page-item active"><a class="page-link">{crrpage}</a></li>
                    '''
                else:
                    page_nav_html += f'''
                    <li class="page-item" onclick="yx_request_data({crrpage})"><a class="page-link">{crrpage}</a></li>
                    '''

            if page == total_page:
                page_nav_html += f'''
                    <li class="page-item disabled">
                        <a class="page-link" aria-label="Next">
                            <span aria-hidden="true">&raquo;</span>
                        </a>
                    </li>                    
                '''
            else:
                page_nav_html += f'''
                    <li class="page-item" onclick="yx_request_data({page + 1})">
                        <a class="page-link" aria-label="Next">
                            <span aria-hidden="true">&raquo;</span>
                        </a>
                    </li>               
                '''

        dataTableBottom_html = f'''   
        <span>总信息条数：{total}，总{total_page}页</span>
        <ul class="pages pagination" data-crrpage="{page}">
            {page_nav_html}                     
        </ul>    
        '''

        return all_datas, total, pages, total_page, table_html, dataTableBottom_html

    def get_yx_game_name_sum_data(self, game_name, start_date, end_date):
        _data = {'game_name': game_name}
        _d1 = game_sfyx_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'platform': game_name}) or []
        _d2 = game_yxtj_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'game_name': game_name}) or []
        for d1 in _d1:
            yx_betting_money = _data.get('yx_betting_money') or 0
            yx_betting_money += d1.get('yx_betting_money')
            fs_money = _data.get('fs_money') or 0
            fs_money += d1.get('fs_money')
            sjsy_money = _data.get('sjsy_money') or 0
            sjsy_money += d1.get('sjsy_money')

            _data.update({
                'yx_betting_money': yx_betting_money,
                'fs_money': fs_money,
                'sjsy_money': sjsy_money,
            })
        for d2 in _d2:
            bs_number = _data.get('bs_number') or 0
            bs_number += d2.get('bs_number')
            betting_money = _data.get('betting_money') or 0
            betting_money += d2.get('betting_money')
            actual_winlose_money = _data.get('actual_winlose_money') or 0
            actual_winlose_money += d2.get('actual_winlose_money')
            xzhy_count = _data.get('xzhy_count') or 0
            xzhy_count += d2.get('xzhy_count')

            _data.update({
                'bs_number': bs_number,
                'betting_money': betting_money,
                'actual_winlose_money': actual_winlose_money,
                'xzhy_count': xzhy_count,
            })

        return _data

    def yx_dbfx(self, game_name, start_date, end_date):
        _data1 = self.get_yx_game_name_sum_data(game_name, start_date, end_date)
        dddy = (start_date - end_date).days
        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        _data2 = self.get_yx_game_name_sum_data(game_name, start_db, end_db)

        fields = [
            'yx_betting_money',
            'fs_money',
            'sjsy_money',
            'bs_number',
            'betting_money',
            'actual_winlose_money',
            'xzhy_count',
        ]
        datas = []
        for filed in fields:
            name = ''
            if filed == 'yx_betting_money':
                name = '有效投注金额'
            if filed == 'fs_money':
                name = '返水金额'
            if filed == 'sjsy_money':
                name = '实际输赢(含返水)'
            if filed == 'bs_number':
                name = '笔数'
            if filed == 'betting_money':
                name = '投注金额'
            if filed == 'actual_winlose_money':
                name = '实际输赢（含退水）'
            if filed == 'xzhy_count':
                name = '下注会员数'

            v1 = _data1.get(filed) or 0
            v2 = _data2.get(filed) or 0
            if v2 == 0 and v1 == 0:
                vv = 0
            elif v2 == 0:
                vv = 100
            else:
                vv = round((v1 - v2) / v2 * 100, 2)
            datas.append({
                'name': name,
                'v1': v1,
                'v2': v2,
                'vv': vv,
            })

        td_html = ''
        for index, d in enumerate(datas):
            td_html += f'''
                    <tr>
                        <td>{ d.get('name') or '' }</td>
                        <td>{ d.get('v1') or 0 }</td>
                        <td>{ d.get('v2') or 0 }</td>
                        <td>{ d.get('vv') }%</td>
                    </tr>                     
            '''

        html = f'''
            <table class="table table-bordered table-hover text-center" style="background-color: #ffffff; font-size: 13px;">
                <tbody>
                    <tr>
                        <td>名称</td>
                        <td>{'%s 至 %s' % (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))}</td>
                        <td>{'%s 至 %s' % (start_db.strftime('%Y-%m-%d'), end_db.strftime('%Y-%m-%d'))}</td>
                        <td>增加百分比</td>
                    </tr>
                    { td_html }
                </tbody>
            </table>        
        '''
        return self.xtjson.json_result(message=html, data={'title': '%s, 对比分析结果' % game_name})


    # 平台报表
    def pt_dbfx_sum_func(self, start_date, end_date):
        _datas = platform_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}})
        _result_dict = {}
        fileds = [
            'app_regist', 'shou_chong', 'xzyxhy_cp', 'xzyxhy_dsfyx', 'yxyh_count', 'yxyh_cp_count', 'yxyh_sfyx_count',
            'cpye_money', 'sfye_money', 'recharge_money', 'withdraw_money', 'touzhu_money', 'tsze_money', 'sjsy_money',
            'agent_odds', 'tuishui_money',
            'jackpot', 'jackBonus', 'sfyx_money', 'sy_money', 'fs_money', 'sfqtfy_money', 'czyh_money', 'hd_money',
            'hb_money', 'yk_mongey',
        ]
        for _d in _datas:
            for fe in fileds:
                _v = _result_dict.get(fe) or 0
                _v += float(_d.get(fe) or 0)
                _result_dict[fe] = _v

        return _result_dict

    def pt_fx_func(self, start_date, end_date):
        _data1 = self.pt_dbfx_sum_func(start_date, end_date)
        dddy = (start_date - end_date).days
        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        _data2 = self.pt_dbfx_sum_func(start_db, end_db)

        fileds = [
            'app_regist', 'shou_chong', 'xzyxhy_cp', 'xzyxhy_dsfyx', 'yxyh_count', 'yxyh_cp_count', 'yxyh_sfyx_count',
            'cpye_money', 'sfye_money', 'recharge_money', 'withdraw_money', 'touzhu_money', 'tsze_money', 'sjsy_money',
            'agent_odds', 'tuishui_money',
            'jackpot', 'jackBonus', 'sfyx_money', 'sy_money', 'fs_money', 'sfqtfy_money', 'czyh_money', 'hd_money',
            'hb_money', 'yk_mongey',
        ]
        result_dict = {}
        datas = []
        for filed in fileds:
            name = ''
            if filed == 'app_regist':
                name = 'APP注册'
            if filed == 'shou_chong':
                name = '首充'
            if filed == 'xzyxhy_cp':
                name = '新增有效会员(彩票)'
            if filed == 'xzyxhy_dsfyx':
                name = '新增有效会员(第三方游戏)'
            if filed == 'yxyh_count':
                name = '有效用户总数'
            if filed == 'yxyh_cp_count':
                name = '有效用户-彩票'
            if filed == 'yxyh_sfyx_count':
                name = '有效用户-第三方'
            if filed == 'cpye_money':
                name = '彩票余额'
            if filed == 'sfye_money':
                name = '三方余额'
            if filed == 'recharge_money':
                name = '充值金额'
            if filed == 'withdraw_money':
                name = '提现金额'
            if filed == 'touzhu_money':
                name = '投注总额'
            if filed == 'tsze_money':
                name = '退水总额'
            if filed == 'sjsy_money':
                name = '实际输赢(含退水)'
            if filed == 'agent_odds':
                name = '代理赔率(金额)'
            if filed == 'tuishui_money':
                name = '退水金额'
            if filed == 'jackpot':
                name = 'jackpot'
            if filed == 'jackBonus':
                name = 'jackBonus'
            if filed == 'sfyx_money':
                name = '三方有效'
            if filed == 'sy_money':
                name = '输赢金额'
            if filed == 'fs_money':
                name = '反水金额'
            if filed == 'sfqtfy_money':
                name = '三方其他费用'
            if filed == 'czyh_money':
                name = '充值优惠&手续费'
            if filed == 'hd_money':
                name = '活动金额'
            if filed == 'hb_money':
                name = '红包金额'
            if filed == 'yk_mongey':
                name = '平台盈亏'
            v1 = _data1.get(filed) or 0
            v2 = _data2.get(filed) or 0
            if v2 == 0 and v1 == 0:
                vv = 0
            elif v2 == 0:
                vv = 100
            else:
                vv = round((v1 - v2) / v2 * 100, 2)
            datas.append({
                'name': name,
                'v1': v1,
                'v2': v2,
                'vv': vv,
            })

        td_html = ''
        for d in datas:
            td_html += f'''
                    <tr>
                        <td>{ d.get('name') or '' }</td>
                        <td>{ d.get('v1') or 0 }</td>
                        <td>{ d.get('v2') or 0 }</td>
                        <td>{ d.get('vv') }%</td>
                    </tr>                     
            '''

        html = f'''
            <table class="table table-bordered table-hover text-center" style="background-color: #ffffff; font-size: 13px;">
                <tbody>
                    <tr>
                        <td>名称</td>
                        <td>{'%s 至 %s' % (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))}</td>
                        <td>{'%s 至 %s' % (start_db.strftime('%Y-%m-%d'), end_db.strftime('%Y-%m-%d'))}</td>
                        <td>增加百分比</td>
                    </tr>
                    { td_html }
                </tbody>
            </table>        
        '''
        return self.xtjson.json_result(message=html, data={'title': '比分析结果'})

    # 总代理报表
    def zdl_get_dl_account_sum_data(self, account, start_date, end_date):
        _data = {'account': account}
        _d1 = agency_caipiao_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []
        _d2 = agency_sanfang_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []
        _d3 = agency_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}) or []

        rech_datas = RechargeTable.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'parent_agent': account}) or []
        hyzj_datas = hy_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'chongzhi_money': {'$gt': 0}, 'parent_agent': account}) or []
        _data['sc_count'] = len(rech_datas)
        _data['czry_count'] = len(hyzj_datas)

        xzzc_datas = CustomerTable.count({'new_time': {'$gte': start_date, '$lte': end_date}, 'agency': account}) or 0
        _data['xzzc_count'] = xzzc_datas

        cus_datas1 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_date, '$lte': end_date}, 'parent_agent': account})
        dl_cu_ls1 = [i.get('account') for i in cus_datas1]
        dddy = (start_date - end_date).days

        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        cus_datas2 = hy_zijing_form_table.find_many({'form_date': {'$gte': start_db, '$lte': end_db}, 'parent_agent': account})
        dl_cu_ls2 = [i.get('account') for i in cus_datas2]
        new_count = 0
        low_count = 0
        for dl in dl_cu_ls1:
            if dl not in dl_cu_ls2:
                new_count += 1
        for dl in dl_cu_ls2:
            if dl not in dl_cu_ls1:
                low_count += 1
        _data['new_count'] = new_count
        _data['low_count'] = low_count

        # 合并投注人数
        ddsls = agency_touzhu_form_table.collection().aggregate([
            {"$match": {'form_date': {'$gte': start_date, '$lte': end_date}, 'account': account}},
            {"$group": {"_id": "$account", "count": {"$sum": '$customer_count'}}},
        ])
        hbtz_ls = {}
        for df in ddsls:
            hbtz_ls[df.get('_id')] = df.get('count') or 0

        _data['hbtz_count'] = hbtz_ls.get(account) or 0

        for d1 in _d1:
            xzhy_count = _data.get('xzhy_count') or 0
            xzhy_count += d1.get('xzhy_count')
            betting_count = _data.get('betting_count') or 0
            betting_count += d1.get('betting_count')
            betting_money = _data.get('betting_money') or 0
            betting_money += d1.get('betting_money')
            actual_winlose_money = _data.get('actual_winlose_money') or 0
            actual_winlose_money += d1.get('actual_winlose_money')
            _data.update({
                'xzhy_count': xzhy_count,
                'betting_count': betting_count,
                'betting_money': betting_money,
                'actual_winlose_money': actual_winlose_money,
            })
        for d2 in _d2:
            bs_number = _data.get('bs_number') or 0
            bs_number += d2.get('bs_number')
            hy_count = _data.get('hy_count') or 0
            hy_count += d2.get('hy_count')
            yx_betting_money = _data.get('yx_betting_money') or 0
            yx_betting_money += d2.get('yx_betting_money')
            hyfs_money = _data.get('hyfs_money') or 0
            hyfs_money += d2.get('hyfs_money')
            sjsy_money = _data.get('sjsy_money') or 0
            sjsy_money += d2.get('sjsy_money')

            _data.update({
                'bs_number': bs_number,
                'hy_count': hy_count,
                'yx_betting_money': yx_betting_money,
                'hyfs_money': hyfs_money,
                'sjsy_money': sjsy_money,
            })
        for d3 in _d3:
            chongzhi_money = _data.get('chongzhi_money') or 0
            chongzhi_money += d3.get('chongzhi_money') or 0
            hdjq_money = _data.get('hdjq_money') or 0
            hdjq_money += d3.get('hdjq_money') or 0
            tixian_money = _data.get('tixian_money') or 0
            tixian_money += d3.get('tixian_money') or 0
            hdkq_money = _data.get('hdkq_money') or 0
            hdkq_money += d3.get('hdkq_money') or 0
            hd_money = _data.get('hd_money') or 0
            hd_money += d3.get('hd_money') or 0
            hb_money = _data.get('hb_money') or 0
            hb_money += d3.get('hb_money') or 0
            czyh_money = _data.get('czyh_money') or 0
            czyh_money += d3.get('czyh_money') or 0

            _data.update({
                'chongzhi_money': chongzhi_money,
                'hdjq_money': hdjq_money,
                'tixian_money': tixian_money,
                'hdkq_money': hdkq_money,
                'hd_money': hd_money,
                'hb_money': hb_money,
                'czyh_money': czyh_money,
            })
        return _data

    def zdl_dbfx_func(self, child_agency, start_date, end_date):
        _data1 = self.zdl_get_dl_account_sum_data(child_agency, start_date, end_date)
        dddy = (start_date - end_date).days
        start_db = start_date + datetime.timedelta(days=dddy)
        end_db = end_date + datetime.timedelta(days=dddy)
        _data2 = self.zdl_get_dl_account_sum_data(child_agency, start_db, end_db)

        fields = [
            'xzhy_count',
            'betting_count',
            'betting_money',
            'actual_winlose_money',
            'bs_number',
            'hy_count',
            'yx_betting_money',
            'hyfs_money',
            'sjsy_money',
            'chongzhi_money',
            'hdjq_money',
            'tixian_money',
            'hdkq_money',
            'hd_money',
            'hb_money',
            'czyh_money',
            'sc_count',
            'czry_count',
            'xzzc_count',
            'hbtz_count',
            'new_count',
            'low_count',
        ]
        datas = []
        for filed in fields:
            name = ''
            if filed == 'xzhy_count':
                name = '彩票下注会员'
            if filed == 'betting_count':
                name = '彩票笔数'
            if filed == 'betting_money':
                name = '投注金额'
            if filed == 'actual_winlose_money':
                name = '彩票输赢'
            if filed == 'bs_number':
                name = '三方笔数'
            if filed == 'hy_count':
                name = '三方下注会员数'
            if filed == 'yx_betting_money':
                name = '三方有效投注金额'
            if filed == 'hyfs_money':
                name = '返水金额'
            if filed == 'sjsy_money':
                name = '三方实际输赢'
            if filed == 'chongzhi_money':
                name = '充值金额'
            if filed == 'hdjq_money':
                name = '后台加钱'
            if filed == 'tixian_money':
                name = '提现金额'
            if filed == 'hdkq_money':
                name = '后台扣钱'
            if filed == 'hd_money':
                name = '活动金额'
            if filed == 'hb_money':
                name = '红包金额'
            if filed == 'czyh_money':
                name = '手续费'
            if filed == 'sc_count':
                name = '首充人数'
            if filed == 'czry_count':
                name = '充值人数'
            if filed == 'xzzc_count':
                name = '新增注册'
            if filed == 'hbtz_count':
                name = '合并投注人数'
            if filed == 'new_count':
                name = '新增人数'
            if filed == 'low_count':
                name = '流失人数'

            v1 = _data1.get(filed) or 0
            v2 = _data2.get(filed) or 0
            if v2 == 0 and v1 == 0:
                vv = 0
            elif v2 == 0:
                vv = 100
            else:
                vv = round((v1 - v2) / v2 * 100, 2)
            datas.append({
                'name': name,
                'v1': v1,
                'v2': v2,
                'vv': vv,
            })

        text1 = f'{start_date.strftime("%m.%d")} - {end_date.strftime("%m.%d")}数据'
        text2 = f'{start_db.strftime("%m.%d")} - {end_db.strftime("%m.%d")}数据'
        return datas, text1, text2

    def zdl_fx_func(self, datas1, uuid, export_folder, filename, text1, text2):
        export_data = ExportDataModel.find_one({'uuid': uuid}) or {}
        if not export_data:
            return
        fields = [
            'account', 'xzhy_count','betting_count','betting_money','actual_winlose_money','bs_number','hy_count','yx_betting_money','hyfs_money','sjsy_money',
            'chongzhi_money','hdjq_money','tixian_money','hdkq_money','hd_money','hb_money','czyh_money','sc_count','czry_count','xzzc_count','hbtz_count','new_count','low_count',
        ]
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 0
            for data in datas1:
                __datas = data.get('datas') or []
                row += 1
                for index, filed in enumerate(fields):
                    if filed == 'account':
                        wa.cell(row=row, column=index+1, value=data.get('account') or '')
                        continue
                    name = ''
                    if filed == 'xzhy_count':
                        name = '彩票下注会员'
                    if filed == 'betting_count':
                        name = '彩票笔数'
                    if filed == 'betting_money':
                        name = '投注金额'
                    if filed == 'actual_winlose_money':
                        name = '彩票输赢'
                    if filed == 'bs_number':
                        name = '三方笔数'
                    if filed == 'hy_count':
                        name = '三方下注会员数'
                    if filed == 'yx_betting_money':
                        name = '三方有效投注金额'
                    if filed == 'hyfs_money':
                        name = '返水金额'
                    if filed == 'sjsy_money':
                        name = '三方实际输赢'
                    if filed == 'chongzhi_money':
                        name = '充值金额'
                    if filed == 'hdjq_money':
                        name = '后台加钱'
                    if filed == 'tixian_money':
                        name = '提现金额'
                    if filed == 'hdkq_money':
                        name = '后台扣钱'
                    if filed == 'hd_money':
                        name = '活动金额'
                    if filed == 'hb_money':
                        name = '红包金额'
                    if filed == 'czyh_money':
                        name = '手续费'
                    if filed == 'sc_count':
                        name = '首充人数'
                    if filed == 'czry_count':
                        name = '充值人数'
                    if filed == 'xzzc_count':
                        name = '新增注册'
                    if filed == 'hbtz_count':
                        name = '合并投注人数'
                    if filed == 'new_count':
                        name = '新增人数'
                    if filed == 'low_count':
                        name = '流失人数'
                    wa.cell(row=row, column=index+1, value=name or '')

                wa.cell(row=row + 1, column=1, value=text1)
                wa.cell(row=row + 2, column=1, value=text2)
                wa.cell(row=row + 3, column=1, value='百分比')

                for index, __d in enumerate(__datas):
                    wa.cell(row=row+1, column=index+2, value=__d.get('v1'))
                    wa.cell(row=row+2, column=index+2, value=__d.get('v2'))
                    wa.cell(row=row+3, column=index+2, value=f'{__d.get("vv")}%')

                row += 4
                wa.cell(row=row, column=1, value='')
                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def zdl_fxff_func(self, main_agency, pt_dateg):
        datas = agencyConfigTable.find_many({'main_agency': main_agency}) or []
        if not datas:
            return self.xtjson.json_params_error('账户不存在！')
        start_date, end_date = PagingCLS.by_silce(pt_dateg)
        zdatas = []
        _text1, _text2 = '', ''
        for _d in datas:
            _crr_datas, _text1, _text2 = self.zdl_dbfx_func(_d.get('child_agency'), start_date, end_date)
            zdatas.append({
                'account': _d.get('child_agency'),
                'datas': _crr_datas
            })
        absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
        export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
        filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
        export_data = {
            'filename': filename,
            'statu': ExportStatu.ongoing,
            'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
            'total': len(datas),
            'out_count': 0,
            'note': f'总代理分析-{main_agency}-{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}'
        }
        duuid = ExportDataModel.insert_one(export_data)

        threading.Thread(target=self.zdl_fx_func, args=(zdatas, duuid, export_folder, filename, _text1, _text2)).start()
        return self.xtjson.json_result()


    def view_get(self):
        self.context['title'] = self.title
        now = datetime.datetime.now() - datetime.timedelta(days=1)
        start_tiem = datetime.datetime(now.year, now.month, now.day, 0, 0, 0)
        end_time = datetime.datetime(now.year, now.month, now.day, 23, 59, 59)
        filter = {'form_date': {'$gte': start_tiem, '$lte': end_time}}
        all_datas, total, pages, total_page, table_html, dataTableBottom_html = self.dl_table_func(filter=filter)
        self.context['pages'] = pages
        self.context['total'] = total
        self.context['total_page'] = total_page
        self.context['table_html'] = table_html
        self.context['dataTableBottom_html'] = dataTableBottom_html
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'get_dl_table_html':
            page = self.request_data.get('page') or 1
            account = self.request_data.get('account') or ''
            form_date = self.request_data.get('form_date') or ''
            filter = {}
            if account:
                filter['account'] = account
            if form_date:
                start_time, end_time = PagingCLS.by_silce(form_date)
                filter['form_date']= {'$gte': start_time, '$lte': end_time}
            all_datas, total, pages, total_page, table_html, dataTableBottom_html = self.dl_table_func(page=int(page), filter=filter)
            search_res = {
                'page': page,
                'account': account,
                'form_date': form_date,
            }
            _result = {
                'search_res': search_res,
                'total': total,
                'total_page': total_page,
                'table_html': table_html,
                'dataTableBottom_html': dataTableBottom_html,
            }
            return self.xtjson.json_result(data=_result)
        if self.action == 'get_dl_db_html':
            _k = current_app.config.get('PROJECT_NAME') + '_' + 'dldbfx_date'
            _v = SiteRedis.get(_k)
            if _v:
                _v = _v.decode()
            return self.get_db_html('dbfx', _v)
        if self.action == 'get_dl_adb_html':
            _k = current_app.config.get('PROJECT_NAME') + '_' + 'dldbfx_date'
            _v = SiteRedis.get(_k)
            if _v:
                _v = _v.decode()
            return self.get_db_html('dbfx', _v)
        if self.action == 'get_hy_db_html':
            _k = current_app.config.get('PROJECT_NAME') + '_' + 'hydbfx_date'
            _v = SiteRedis.get(_k)
            if _v:
                _v = _v.decode()
            return self.get_db_html('hy_dbfx', _v)
        if self.action == 'get_yx_db_html':
            _k = current_app.config.get('PROJECT_NAME') + '_' + 'yxdbfx_date'
            _v = SiteRedis.get(_k)
            if _v:
                _v = _v.decode()
            return self.get_db_html('yx_dbfx', _v)
        if self.action == 'dbfx':
            acount = self.request_data.get('acount')
            db_date = self.request_data.get('db_date')
            if not acount or not db_date:
                return self.xtjson.json_params_error()
            _k = current_app.config.get('PROJECT_NAME') + '_' + 'dldbfx_date'
            SiteRedis.set(_k, db_date)
            start_date, end_date = PagingCLS.by_silce(db_date)
            return self.dbfx_func(acount, start_date, end_date)
        if self.action == 'get_hy_table_html':
            page = self.request_data.get('page') or 1
            account = self.request_data.get('account') or ''
            form_date = self.request_data.get('form_date') or ''
            parent_agent = self.request_data.get('parent_agent') or ''
            filter = {}
            if account:
                filter['account'] = account
            if parent_agent:
                filter['parent_agent'] = parent_agent
            if form_date:
                start_time, end_time = PagingCLS.by_silce(form_date)
                filter['form_date']= {'$gte': start_time, '$lte': end_time}
            all_datas, total, pages, total_page, table_html, dataTableBottom_html = self.hy_table_func(page=int(page), filter=filter)
            search_res = {
                'page': page,
                'account': account,
                'form_date': form_date,
                'parent_agent': parent_agent,
            }
            _result = {
                'search_res': search_res,
                'total': total,
                'total_page': total_page,
                'table_html': table_html,
                'dataTableBottom_html': dataTableBottom_html,
            }
            return self.xtjson.json_result(data=_result)
        if self.action == 'hy_dbfx':
            acount = self.request_data.get('acount')
            db_date = self.request_data.get('db_date')
            if not acount or not db_date:
                return self.xtjson.json_params_error()
            _k = current_app.config.get('PROJECT_NAME') + '_' + 'hydbfx_date'
            SiteRedis.set(_k, db_date)
            start_date, end_date = PagingCLS.by_silce(db_date)
            return self.hy_dbfx_func(acount, start_date, end_date)
        if self.action == 'pt_fx':
            pt_date = self.request_data.get('pt_date')
            if not pt_date:
                return self.xtjson.json_params_error()
            start_date, end_date = PagingCLS.by_silce(pt_date)
            return self.pt_fx_func(start_date, end_date)
        if self.action == 'export_ls':
            return self.export_ls_func()
        if self.action == 'get_yx_table_html':
            page = self.request_data.get('page') or 1
            game_name = self.request_data.get('game_name') or ''
            form_date = self.request_data.get('form_date') or ''
            filter = {}
            if game_name:
                filter['game_name'] = game_name
            if form_date:
                start_time, end_time = PagingCLS.by_silce(form_date)
                filter['form_date']= {'$gte': start_time, '$lte': end_time}
            all_datas, total, pages, total_page, table_html, dataTableBottom_html = self.yx_table_func(page=int(page), filter=filter)
            search_res = {
                'page': page,
                'game_name': game_name,
                'form_date': form_date,
            }
            _result = {
                'search_res': search_res,
                'total': total,
                'total_page': total_page,
                'table_html': table_html,
                'dataTableBottom_html': dataTableBottom_html,
            }
            return self.xtjson.json_result(data=_result)
        if self.action == 'yx_dbfx':
            game_name = self.request_data.get('game_name')
            db_date = self.request_data.get('db_date')
            if not game_name or not db_date:
                return self.xtjson.json_params_error()
            _k = current_app.config.get('PROJECT_NAME') + '_' + 'yxdbfx_date'
            SiteRedis.set(_k, db_date)
            start_date, end_date = PagingCLS.by_silce(db_date)
            return self.yx_dbfx(game_name, start_date, end_date)
        if self.action == 'get_hy_tzfx_html':
            return self.get_hy_tzfx_html()
        # 总代理报表
        if self.action == 'zdl_fx':
            pt_dateg = self.request_data.get('pt_dateg')
            main_agency = self.request_data.get('main_agency')
            if not main_agency or not main_agency.strip()or not pt_dateg:
                return self.xtjson.json_params_error()
            return self.zdl_fxff_func(main_agency, pt_dateg)


class agencyConfigView(DataFormViewBase):
    title = '总代理配置'
    show_menu = False
    add_url_rules = [['/agencyConfig', 'agencyConfig']]
    template = 'data_form/agencyConfig.html'
    MCLS = agencyConfigTable
    per_page = 30

    def add_form_html(self):
        html = f'''
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>总代理：</span></div>
                    <input type="text" class="form-control " id="main_agency" placeholder="总代理" value="">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>子代理：</span></div>
                    <textarea class="form-control" id="child_agency" rows="7" aria-label="With textarea" placeholder="子代理,一行一个"></textarea>
                </div>                
                <span class="btn btn-primary swal2-styled" onclick="post_agencyConfig()" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>        
            '''
        return self.xtjson.json_result(message=html)

    def add_agencyConfig(self):
        main_agency = self.request_data.get('main_agency')
        child_agency = self.request_data.get('child_agency')
        if not main_agency or not child_agency:
            return self.xtjson.json_params_error()
        main_agency = main_agency.strip()
        child_agencys = child_agency.strip().split('\n')
        for chdata in child_agencys:
            if self.MCLS.find_one({'main_agency': main_agency, 'child_agency': chdata}):
                continue
            _d = {'main_agency': main_agency, 'child_agency': chdata}
            self.MCLS.insert_one(_d)
        return self.xtjson.json_result()

    def get_context(self):
        return {'format_money': self.format_money}

    def post_other_way(self):
        if self.action == 'del_all':
            self.MCLS.delete_many({})
            return self.xtjson.json_result()
        if self.action == 'add_form_html':
            return self.add_form_html()
        if self.action == 'add_agencyConfig':
            return self.add_agencyConfig()

    def post_data_other_way(self):
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()


