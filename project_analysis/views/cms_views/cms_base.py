# -*- coding: utf-8 -*-
import csv, os, datetime
from threading import Thread
from openpyxl import Workbook
from flask import views, request, session, redirect, url_for, abort, render_template, current_app
from . import bp
from common_utils import xtjson
from common_utils.upload import UploadCls
from models.cms_user import CmsUserModel
from common_utils.utils_funcs import PagingCLS
from ..view_func import add_admin_log, cms_risk_control
from common_utils.utils_funcs import generate_filename
from constants import CMS_USER_SESSION_KEY, OnClick, SITE_CONFIG_CACHE, CMS_MENUS_GROUPS, EventType, ExportStatu, EXPORT_FOLDER, IMAGES_TYPES, ASSETS_FOLDER, PermissionType
from models.customer_table import ExportDataModel


@bp.before_request
def site_cms_before_request():
    res = cms_risk_control()
    if res:
        return res


class CmsViewBase(views.MethodView):

    title = ''
    template = ''
    endpoint = ''
    xtjson = xtjson
    permission_map = {}
    MCLS = None
    add_url_rules = [[]]

    def __init_data(self):
        self.user_uuid = session.get(CMS_USER_SESSION_KEY)
        self.current_admin_user = CmsUserModel.query_one({'uuid': self.user_uuid})
        self.is_superdamin = False
        if self.current_admin_user:
            self.is_superdamin = self.current_admin_user.is_superadmin
        self.EventType = EventType
        self.data_from = {}
        self.filter_dict = {}
        self.context = {}
        self.project_name = current_app.config.get('PROJECT_NAME')
        self.project_static_folder = os.path.join(current_app._static_folder, self.project_name)
        self.context['current_admin_user'] = self.current_admin_user
        self.context['site_data'] = SITE_CONFIG_CACHE
        self.context['load_menus'] = self.load_menus
        self.context['cms_menu_groups'] = CMS_MENUS_GROUPS
        self.context['FIELDS'] = self.MCLS.fields()
        self.context['is_superdamin'] = self.is_superdamin

    @classmethod
    def html_StringField(cls, db_field, field_cls, data_dict={}):
        html = '<div class="input-group mb-3">'
        html += '<div class="input-group-prepend">'
        html += '<span class="input-group-text">'
        if not field_cls.nullable:
            html += f'<span class="text-danger">*</span>{field_cls.field_name}：</span></div>'
        else:
            html += f'{field_cls.field_name}：</span></div>'
        if data_dict.get(db_field):
            html += u'<input type="text" class="form-control" id="%s" placeholder="%s" value="%s">' % (db_field, field_cls.placeholder, data_dict.get(db_field))
        else:
            html += u'<input type="text" class="form-control " id="%s" placeholder="%s">' % (db_field, field_cls.placeholder)
        html += u'</div>'
        return html

    @classmethod
    def html_BooleanField(cls, db_field, field_cls, data_dict={}):
        html = u'<div class="input-group mb-3">'
        if not field_cls.nullable:
            html += u'<div class="input-group-prepend"><span class="input-group-text"><span style="color: #c00">*</span>%s：</span></div>' % field_cls.field_name
        else:
            html += u'<div class="input-group-prepend"><span class="input-group-text">%s:</span></div>' % field_cls.field_name
        html += u'<select id="%s" class="form-control">' % db_field
        for cb in field_cls.choices:
            if data_dict.get(db_field) == cb[0]:
                html += u'<option value="%s" selected>%s</option>' % (cb[0], cb[1])
                continue
            html += u'<option value="%s">%s</option>' % (cb[0], cb[1])
        html += u'</select>'
        html += u'</div>'
        return html

    @classmethod
    def html_DictField(cls, db_field, field_cls, data_dict={}):
        html = u'<div class="input-group mb-3">'
        if not field_cls.nullable:
            html += u'<div class="input-group-prepend"><span class="input-group-text"><span style="color: #c00">*</span> %s:</span></div>' % field_cls.field_name
        else:
            html += u'<div class="input-group-prepend"><span class="input-group-text">%s:</span></div>' % field_cls.field_name
        html += u'<select id="%s" class="form-control">' % db_field
        html += f'<option value="">{field_cls.placeholder or field_cls.field_name}</option>'
        for dt in field_cls.dict_cls.name_arr:
            if data_dict.get(db_field) == dt:
                html += u'<option value="%s" selected>%s</option>' % (dt, field_cls.dict_cls.name_dict.get(dt))
                continue
            html += u'<option value="%s">%s</option>' % (dt, field_cls.dict_cls.name_dict.get(dt))
        html += u'</select>'
        html += u'</div>'
        return html

    @classmethod
    def html_DateTimeField(cls, db_field, field_cls, data_dict={}):
        html = '<div class="input-group mb-3">'
        html += '<div class="input-group-prepend">'
        html += '<span class="input-group-text">'
        if not field_cls.nullable:
            html += f'<span class="text-danger">*</span>{field_cls.field_name}：</span></div>'
        else:
            html += f'{field_cls.field_name}：</span></div>'
        if data_dict.get(db_field):
            html += f'<input type="text" class="form-control" onmouseover="reload_picker_time($(this))" id="%s" placeholder="%s" value="%s" readonly>' % (db_field, field_cls.placeholder, data_dict.get(db_field))
        else:
            html += f'<input type="text" class="form-control" onmouseover="reload_picker_time($(this))" id="%s" placeholder="%s" readonly>' % (db_field, field_cls.placeholder)
        html += '</div>'
        return html

    @classmethod
    def html_RelationField(cls, db_field, field_cls, data_dict={}):
        html = '<div class="input-group mb-3">'
        html += '<div class="input-group-prepend">'
        html += '<span class="input-group-text">'
        if not field_cls.nullable:
            html += f'<span class="text-danger">*</span>{field_cls.field_name}：</span></div>'
        else:
            html += f'{field_cls.field_name}：</span></div>'
        html += '<select id="%s" class="form-control">' % db_field
        html += '<option value="">%s</option>' % field_cls.field_name
        relation_data = field_cls.relation_datas()
        for dt in relation_data:
            if data_dict.get(db_field) == dt.get('uuid'):
                html += '<option value="%s" selected>%s</option>'%(dt.get('uuid'), dt.get(field_cls.relation_show_field))
                continue
            html += '<option value="%s">%s</option>'%(dt.get('uuid'), dt.get(field_cls.relation_show_field))
        html += '</select>'
        html += '</div>'
        return html

    def load_menus(self, datas=[]):
        if not datas:
            return []
        p_menus = []
        for menu in datas:
            if self.check_permission('%s&%s' % (menu.get('endpoint'), PermissionType.ACCESS)):
                p_menus.append(menu)
        return p_menus

    def check_login(self):
        cur_user = {}
        # if CMS_USER_SESSION_KEY in session:
        #     cur_user = session[CMS_USER_SESSION_KEY]

        # print("CurUser:", cur_user)
        
        # self.current_admin_user = cur_user

        if not self.current_admin_user:
            self.addcmslog_func(EventType.ACCESS, '未登录！跳转登录页面')
            return redirect(url_for('admin.cms_login'))

        # if not self.current_admin_user.statu:
        #     session.pop(CMS_USER_SESSION_KEY)
        #     return abort(404)
        # if not self.check_permission('%s&%s'%(self.endpoint, PermissionType.ACCESS)):
        #     return self.no_permission()
        # return render_template('fenXi/index.html')

    def check_permission(self, code):
        if self.is_superdamin:
            return True
        if not self.current_admin_user or not code:
            return False
        return self.current_admin_user.has_permission(code)

    def check_superdamin(self, pers):
        if 'superadmin' in pers:
            return True
        return

    def addcmslog_func(self, event_type, msg):
        if self.current_admin_user:
            add_admin_log(event_type, msg, admin_uuid=self.current_admin_user.uuid)
        else:
            add_admin_log(event_type, msg)

    def search_func(self, FIELDS):
        """get数据搜索处理"""
        s_filter_dict, s_context_res = {}, {}
        if hasattr(self.MCLS, 'field_search'):
            field_search = getattr(self.MCLS, 'field_search')()
            for db_field in field_search:
                col_value = request.args.get(db_field)
                if col_value is None:
                    continue
                s_context_res[db_field] = col_value
                if col_value and col_value.strip():
                    col_value = col_value.strip()
                    field_cls = FIELDS.get(db_field)
                    if not field_cls:
                        return  False, '%s: 无处理属性!' % db_field
                    if field_cls.field_type == 'UUIDField':
                        s_filter_dict[db_field] = col_value
                    elif field_cls.field_type == 'IDField':
                        s_filter_dict[db_field] = int(col_value)
                    else:
                        statu, res = field_cls.search_validate(col_value)
                        if not statu:
                            return False, res
                        s_filter_dict[db_field] = res
        return True, [s_filter_dict, s_context_res]

    def export_xlsx_func(self, uuid, export_folder, filename, datas, headers, filead_datas=[]):
        if not os.path.exists(export_folder):
            os.makedirs(export_folder)

        export_data = ExportDataModel.find_one({'uuid': uuid}) or {}
        if not export_data:
            return

        try:
            crr_count = 0
            wb = Workbook()
            wa = wb.active
            row = 1
            for h in range(len(headers)):
                wa.cell(row=row, column=h + 1, value=headers[h])

            for data in datas:
                row += 1
                for index, filed in enumerate(filead_datas):
                    _v = data.get(filed.get('db_filed'))
                    if filed.get('type') == 'int':
                        wa.cell(row=row, column=index+1, value=int(_v or 0))
                    elif filed.get('type') == 'float':
                        wa.cell(row=row, column=index+1, value=float(_v or 0))
                    elif filed.get('type') == 'datetime':
                        if _v and isinstance(_v, datetime.datetime):
                            wa.cell(row=row, column=index+1, value=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        wa.cell(row=row, column=index + 1, value=_v or '')
                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def editorUploadFile(self, foldername='article/images'):
        """编辑器上传文件处理"""
        if not self.check_permission('%s&%s'%(self.endpoint, PermissionType.UPLOAD_DATA)):
            data = { "uploaded": 0, "error": { "message": '无文件操作权限!'}}
            return xtjson.json_result(**data)
        upload = UploadCls()
        upload.static_folder = self.project_static_folder
        upload.uploaddir = ASSETS_FOLDER
        upload.foldername = foldername
        upload.limit_types = IMAGES_TYPES
        statu, msg = upload.upload_file_func()
        if statu:
            filename = msg.rsplit('/', 1)[-1]
            result_dict = {"uploaded": 1, "fileName": filename, "url": msg}
        else:
            result_dict = {"uploaded": 0, "error": { "message": msg}}
        return xtjson.json_result(**result_dict)

    def post_upload_picture(self, picture_name='', foldername='article/images', limit_types=IMAGES_TYPES, limit_size=100):
        """
        图片上传处理
        :param picture_name: 图片名称
        :param uploaddir: 上传到的文件夹名称(1级)
        :param foldername:  上传到的文件父级文件夹名称(2级)
        :param limit_size: 文件大小限制(KB)
        :param limit_types:
        :return:
        """
        if not self.check_permission('%s&%s'%(self.endpoint, PermissionType.UPLOAD_DATA)):
            return self.no_permission()
        upload = UploadCls()
        upload.static_folder = self.project_static_folder
        upload.uploaddir = ASSETS_FOLDER
        upload.foldername = foldername
        upload.limit_types = limit_types
        upload.filename = picture_name
        statu, msg = upload.upload_file_func()
        return statu, msg

    def post_file_upload(self, filename='', foldername='', limit_types=IMAGES_TYPES):
        """
        文件上传处理
        :param filename: 文件名
        :param uploaddir: 上传到的文件夹名称(1级)
        :param foldername:  上传到的文件父级文件夹名称(2级)
        :param limmit_types:  文件类型限制
        """
        if not self.check_permission('%s_%s'%(self.endpoint, PermissionType.UPLOAD_DATA)):
            return self.no_permission()
        upload = UploadCls()
        upload.static_folder = self.project_static_folder
        upload.uploaddir = ASSETS_FOLDER
        upload.foldername = foldername
        upload.limit_types = limit_types
        upload.filename = filename
        return upload.upload_file_func()

    def has_permission_func(self):
        """权限检测"""
        operation_type = ('_add_', '_del_', '_edit_', '_out_', '_upload_')
        if not self.check_permission('%s_%s'%(self.endpoint, PermissionType.ACCESS)):
            return
        for oper in operation_type:
            if self.action.startswith(oper):
                if oper == '_add_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.ADD)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_del_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.DELETE)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_edit_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.EDIT)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_out_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.EXPORT_DATA)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_upload_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.UPLOAD_DATA)
                    if not self.check_permission(crr_p):
                        return
        return True

    def no_permission(self):
        return self.xtjson.json_params_error('无操作权限!')

    def get_other_way(self):
        return

    def get_context(self):
        """获取context内容"""
        return {}

    def get_filter_dict(self):
        """获取搜索参数"""
        return {}

    def post_data_other_way(self):
        return

    def post_other_way(self):
        return

    def view_get(self, *args, **kwargs):
        return self.xtjson.json_result()

    def view_post(self, *args, **kwargs):
        return self.xtjson.json_result()

    def get(self, *args, **kwargs):
        self.__init_data()
        res = self.check_login()
        if res:
            return res
        add_admin_log(EventType.ACCESS)
        return self.view_get(*args, **kwargs)

    def post(self, *args, **kwargs):
        self.__init_data()
        res = self.check_login()
        if res:
            return res
        return self.view_post(*args, **kwargs)


class CmsFormViewBase(CmsViewBase):
    title = ''
    template = ''
    show_menu = True
    MCLS = CmsUserModel

    def post_edit_data(self, data_form):
        self.data_dict.update(data_form)
        self.MCLS.save(self.data_dict)
        return xtjson.json_result()

    def checkPermission(self, code):
        if code in self.current_admin_user.permissions or self.current_admin_user.is_superadmin:
            return
        return xtjson.json_params_error('无权限！')

    def view_post(self, *args, **kwargs):
        self.kwargs = kwargs or {}
        self.request_data = request.form
        self.action = self.request_data.get('action')
        self.data_uuid = self.request_data.get('data_uuid')
        self.data_value = self.request_data.get('data_value')
        if request.args.get('CKEditorFuncNum'):
            return self.editorUploadFile()
        res = self.post_other_way()
        if res:
            return res
        self.data_dict = self.MCLS.find_one({'uuid': self.data_uuid})
        if not self.data_dict:
            return xtjson.json_params_error('数据不存在!')
        if self.action == '_edit_form_data':
            for db_field in self.MCLS.edit_field_sort():
                field_cls = self.MCLS.fields().get(db_field)
                v = self.request_data.get(db_field)
                statu, res = field_cls.validate(v)
                if not statu:
                    return xtjson.json_params_error(res)
                self.data_from[db_field] = res
            return self.post_edit_data(self.data_from)
        if self.action == '_del_':
            self.MCLS.delete_one(self.data_dict)
            return xtjson.json_result()
        res = self.post_data_other_way()
        if res:
            return res
        return xtjson.json_params_error('操作错误!')


class CmsTableViewBase(CmsViewBase):
    per_page = 100
    show_menu = True
    MCLS = CmsUserModel
    multi_select = False
    sort = [['_create_time', -1]]

    def out_data_html(self, file_type):
        out_filename = generate_filename()
        html = ''
        html += '<table class="table table-bordered table-hover text-center" style="background-color: #ffffff">'
        html += '<thead class="thead-light"><tr>'
        html += '<th>字段名称</th>'
        html += '<th>选择导出字段</th>'
        html += '</tr></thead>'
        html += '<tbody>'
        for db_field in self.MCLS.fields().keys():
            field_Cls = self.MCLS.fields().get(db_field)
            if field_Cls:
                html += '<tr>'
                html += f'<td>{ field_Cls.field_name }</td>'
                html += f'<td><input type="checkbox" class="form-check-input search_out_checkbox" checked data-type="{db_field}" style="margin: 0;position: relative; top: 3px;"></td>'
                html += '</tr>'
        html += '</tbody></table>'
        html += f"""
            <div class="input-group mb-3">
                <div class="input-group-prepend">
                    <span class="input-group-text">导出文件名:</span>
                </div>
                <input type="text" class="form-control" id="out_filename" value="{ out_filename }" placeholder="导出文件名">
                <div class="input-group-append">
                    <span class="input-group-text">{file_type}</span>
                </div>
            </div>        
            <div class="input-group mb-3">
                <div class="input-group-prepend">
                    <span class="input-group-text">导出数量:</span>
                </div>
                <input type="text" class="form-control" id="out_count" value="" placeholder="导出数量,默认当前数据全部导出">
            </div>            
        """
        return xtjson.json_result(message=html)

    def del_search_data(self, _search_filter_dict):
        """删除搜索数据"""
        for _data in self.MCLS.find_many(_search_filter_dict):
            self.MCLS.delete_one({'uuid': _data.get('uuid')})

    def del_multi_select(self, select_uuids):
        """删除多选"""
        for uuid in select_uuids:
            self.MCLS.delete_one({'uuid': uuid})

    def process_data(self, datas):
        """数据处理"""
        return datas

    def post_data_del_all(self, filter_dict={}):
        self.MCLS.delete_many(filter_dict)
        return xtjson.json_result()

    def post_data_del(self):
        self.MCLS.delete_one(self.data_dict)
        return xtjson.json_result()

    def post_add_data(self, data_form):
        self.MCLS.insert_one(data_form)
        return xtjson.json_result()

    def post_edit_data(self, data_form):
        self.data_dict.update(data_form)
        self.MCLS.save(self.data_dict)
        return xtjson.json_result()

    def view_get(self):
        res = self.get_other_way()
        if res:
            return res
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        self.context['multi_select'] = self.multi_select
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)
        for data in all_datas:
            for db_field in self.MCLS.field_sort():
                field_cls = fields.get(db_field)
                if field_cls.field_type == 'RelationField':
                    relation_data = field_cls.relation_data(data.get(db_field))
                    if relation_data:
                        data[db_field] = relation_data.get(field_cls.relation_show_field) or ''
        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = self.process_data(all_datas)
        self.context['pagination'] = pagination
        self.context['context_res'] = context_res
        self.context['OnClick'] = OnClick
        self.context['add_field_sort'] = self.MCLS.add_field_sort()
        self.context['edit_field_sort'] = self.MCLS.edit_field_sort()
        self.context['field_search'] = self.MCLS.field_search()
        self.context['field_sort'] = self.MCLS.field_sort()
        return render_template(self.template, **self.context)

    def view_post(self):
        self.request_data = request.form
        self.action = self.request_data.get('action')
        self.data_uuid = self.request_data.get('data_uuid')
        self.data_value = self.request_data.get('data_value')
        if not self.action:
            return xtjson.json_params_error('操作错误!')
        if not self.has_permission_func():
            return self.no_permission()
        res = self.post_other_way()
        if res:
            return res
        if self.action == '_del_all_data':
            Thread(target=self.post_data_del_all).start()
            return self.xtjson.json_result(message='后台删除中...')
        if self.action == '_del_multi_select':
            select_uuids = self.request_data.getlist('select_uuids[]')
            if not select_uuids:
                return self.xtjson.json_params_error('未选择数据!')
            Thread(target=self.del_multi_select, args=(select_uuids,)).start()
            return self.xtjson.json_result(message='后台删除中...')
        if self.action == '_del_search_data':
            statu, _msg = self.search_func(self.MCLS.fields())
            if not statu:
                return self.xtjson.json_params_error(_msg)
            if not _msg:
                return self.xtjson.json_params_error('')
            _search_filter_dict, _ = _msg
            if not _search_filter_dict:
                return self.xtjson.json_params_error('未搜索数据！')
            Thread(target=self.del_search_data, args=(_search_filter_dict,)).start()
            return self.xtjson.json_result(message='后台删除中，稍后刷新查看...')
        if self.action == '_out_execl_html':
            return self.out_data_html('.xlsx')
        if self.action == '_out_csv_html':
            return self.out_data_html('.csv')
        if self.action == '_add_form_html':
            if not self.MCLS.add_field_sort():
                return xtjson.json_params_error('无处理方式!')
            html = ''
            for db_field in self.MCLS.add_field_sort():
                field_cls = self.MCLS.fields().get(db_field)
                if hasattr(self, 'html_%s'%(field_cls.field_type)):
                    html += getattr(self, 'html_%s'%(field_cls.field_type))(db_field, field_cls)
                else:
                    html += self.html_StringField(db_field, field_cls)
            return self.xtjson.json_result(message=html)
        if self.action == '_add_form_data':
            for db_field in self.MCLS.add_field_sort():
                field_cls = self.MCLS.fields().get(db_field)
                v = self.request_data.get(db_field)
                statu, res = field_cls.validate(v)
                if not statu:
                    return xtjson.json_params_error(res)
                self.data_from[db_field] = res
            return self.post_add_data(self.data_from)
        self.data_dict = self.MCLS.find_one({'uuid': self.data_uuid})
        if not self.data_dict:
            return xtjson.json_params_error('数据不存在!')
        if self.action == '_edit_form_html':
            if not self.MCLS.edit_field_sort():
                return xtjson.json_params_error('无处理方式!')
            html = ''
            for db_field in self.MCLS.edit_field_sort():
                field_cls = self.MCLS.fields().get(db_field)
                if hasattr(self, 'html_%s'%(field_cls.field_type)):
                    html += getattr(self, 'html_%s'%(field_cls.field_type))(db_field, field_cls, self.data_dict)
                else:
                    html += self.html_StringField(db_field, field_cls, self.data_dict)
            return self.xtjson.json_result(message=html)
        if self.action == '_edit_form_data':
            if not self.MCLS.edit_field_sort():
                return xtjson.json_params_error('无处理方式!')
            for db_field in self.MCLS.edit_field_sort():
                field_cls = self.MCLS.fields().get(db_field)
                v = self.request_data.get(db_field)
                statu, res = field_cls.validate(v)
                if not statu:
                    return xtjson.json_params_error(res)
                self.data_from[db_field] = res
            return self.post_edit_data(self.data_from)
        res = self.post_data_other_way()
        if self.action == '_del_':
            return self.post_data_del()
        if self.action == '_edit_note':
            if 'note' in self.MCLS.fields().keys():
                if self.data_value.lower() == '_yes_':
                    self.data_dict['note'] = ''
                else:
                    self.data_dict['note'] = self.data_value.strip()
                self.MCLS.save(self.data_dict)
                return xtjson.json_result()
        if self.action == '_edit_statu':
            if 'statu' in self.MCLS.fields().keys():
                if self.data_dict.get('statu'):
                    self.data_dict['statu'] = 0
                else:
                    self.data_dict['statu'] = 1
                self.MCLS.save(self.data_dict)
                return xtjson.json_result()
        if res:
            return res
        return xtjson.json_params_error('操作错误!')

