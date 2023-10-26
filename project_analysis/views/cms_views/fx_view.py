# -*- coding: utf-8 -*-
import datetime, os, shortuuid, threading, multiprocessing, csv, time, random
from openpyxl import Workbook
from flask import request, render_template, current_app, abort
from .cms_base import CmsFormViewBase
from models.fenxi_table import ZhuDanTable, ImportLogTable, BettingDataTable, ChongZhiFenXiTable, FtpTable
from common_utils.utils_funcs import PagingCLS
from constants import PermissionCls, ASSETS_FOLDER, EXPORT_FOLDER, ExportStatu
from models.cms_user import CmsUserModel
from models.fenxi_table import SiteSettingTable
from models.customer_table import ExportDataModel
from views.view_func import zd_fenxi_func, fenxi_task_func, fcFenXiCsv_fucn, get_new_zhudan_info_func
from common_utils.lqredis import SiteRedis
from modules.ftp_cls import FtpCls
from werkzeug.security import generate_password_hash, check_password_hash



class ZhuDanAnalysis(CmsFormViewBase):
    title = '注单分析'
    show_menu = False
    add_url_rules = [['/analysis/', 'analysis']]
    template = 'fenXi/analysis.html'
    MCLS = ZhuDanTable

    def view_get(self):
        self.context['title'] = self.title
        cs_datas = ZhuDanTable.distinct('game_name',{})
        self.context['cs_datas'] =cs_datas
        return render_template(self.template, **self.context)

    def post_other_way(self):
        date_number = self.request_data.get('date_number')
        youxi = self.request_data.get('youxi')
        if self.action == 'fenXi':
            if not date_number:
                return self.xtjson.json_result('缺少期号!')
            try:
                daten = datetime.datetime.strptime(date_number, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('期号格式错误！')
            if not self.MCLS.find_one({'date_number': daten.strftime('%d/%m/%Y'), 'game_name': youxi}):
                return self.xtjson.json_params_error('该日期下，没有当前游戏数据！')

            _html = ''
            new_wfxz_zongE_ls = []

            _crrKey = 'cs_0'
            res = fenxi_task_func(youxi, daten.strftime('%d/%m/%Y'), _crrKey)
            _html += res.get('res_html')
            new_wfxz_zongE_ls.append(res.get('result_data').get('new_wfxz_zongE'))
            _data = {
                'new_wfxz_zongE_ls': new_wfxz_zongE_ls,
                'html': _html,
            }
            return self.xtjson.json_result(message=_data)
        if self.action == 'getYouXi':
            try:
                daten = datetime.datetime.strptime(date_number, '%Y-%m-%d')
            except:
                return self.xtjson.json_params_error('期号格式错误！')
            cs_datas = ZhuDanTable.distinct('game_name', {'date_number': daten.strftime('%d/%m/%Y')})
            dhtml = '<option value="">选择游戏</option>'
            for cs_data in cs_datas:
                dhtml += f'<option value="{cs_data}">{cs_data}</option>'
            return self.xtjson.json_result(message=dhtml)



class ZhuDanList(CmsFormViewBase):
    MCLS = ZhuDanTable
    title = '注单数据管理'
    add_url_rules = [['/zhuDan/', 'zhuDanManage']]
    template = 'fenXi/zhuDan.html'
    per_page = 20
    sort = [['create_time', -1]]

    def fenXi_csv(self, filePath):
        name_dict = {
            '订单号': 'order_number',
            '会员': 'member_account',
            '下注时间': 'create_time',
            '游戏': 'game_name',
            '期号': 'date_number',
            '玩法分类': 'play_type',
            '玩法名称': 'play_name',
            '注单信息': 'zhudan_info',
            '下注金额': 'money',
            '会员输赢': 'win_lose',
        }

        csv_data = []
        try:
            with open(filePath, 'r', encoding='gbk') as fp:
                csv_data = list(csv.DictReader(fp))
        except:
            try:
                with open(filePath, 'r', encoding='utf8') as fp:
                    csv_data = list(csv.DictReader(fp))
            except:
                with open(filePath, 'r', encoding='GB2312') as fp:
                    csv_data = list(csv.DictReader(fp))


        pool = multiprocessing.Pool(15)
        multi_result = []
        while csv_data:
            crr_datas = csv_data[:5000]
            if not crr_datas:
                break
            del csv_data[:5000]
            multi_result.append(pool.apply_async(func=zd_fenxi_func, args=(crr_datas, name_dict)))
        pool.close()
        pool.join()
        zd_fenxi_func(csv_data, name_dict)

        total_datas = []
        for m in multi_result:
            total_datas += m.get(0)

        return total_datas

    def del_all(self):
        self.MCLS.delete_many({})

    def format_datetime(self, data):
        if isinstance(data, datetime.datetime):
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    def import_data(self, log_uuid, datas):
        for da in datas:
            da['log_uuid'] = log_uuid
            self.MCLS.update_one({'order_number': da.get('order_number')}, {'$set': da},upsert=True)

    def view_get(self):
        if 'zhudanData' not in self.current_admin_user.permissions and not self.current_admin_user.is_superadmin:
            return abort(404)

        self.context['title'] = self.title
        self.context['format_datetime'] = self.format_datetime
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'importData':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('csv'):
                return self.xtjson.json_params_error('文件格式错误，只支持CSV文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/importFile/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)

            datas = self.fenXi_csv(import_folder + new_filename + fext)
            if not datas:
                return self.xtjson.json_params_error('文件无数据！')

            static_path = (import_folder + new_filename + fext).replace(current_app.root_path, '')
            log_uuid = shortuuid.uuid()
            _log = {
                'uuid': log_uuid,
                'file_name': fileobj.filename,
                'file_path': static_path,
                'data_count': len(datas),
                'create_time': datetime.datetime.now()
            }
            ImportLogTable.insert_one(_log)

            threading.Thread(target=self.import_data, args=(log_uuid, datas)).start()
            return self.xtjson.json_result(message='数据提交成功，后台储存中，请稍后刷新查看！')
        if self.action == 'del_all':
            threading.Thread(target=self.del_all).start()
            return self.xtjson.json_result(message='删除中， 请稍后刷新查看！')



class ImportLogManage(CmsFormViewBase):
    MCLS = ImportLogTable
    title = '注单数据导入日志管理'
    add_url_rules = [['/importLog/', 'importLog']]
    template = 'fenXi/importLog.html'
    per_page = 20
    sort = [['create_time', -1]]

    def del_data(self):
        ZhuDanTable.delete_many({'log_uuid': self.data_uuid})
        self.MCLS.delete_one({'uuid': self.data_uuid})
        return

    def format_datetime(self, data):
        if isinstance(data, datetime.datetime):
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    def view_get(self):
        if 'exportLog' not in self.current_admin_user.permissions and not self.current_admin_user.is_superadmin:
            return abort(404)

        self.context['title'] = self.title
        self.context['format_datetime'] = self.format_datetime
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_data_other_way(self):
        if self.action == 'del':
            threading.Thread(target=self.del_data).start()
            return self.xtjson.json_result(message='数据删除中， 请稍后刷新查看！')

    def post_other_way(self):
        if self.action == 'del_all':
            for d in self.MCLS.find_many({}):
                ZhuDanTable.find_many({'log_uuid': d.get('uuid')})
                self.MCLS.delete_one({'uuid': d.get('uuid')})
            return self.xtjson.json_result(message='删除成功！')



class SiteConfigView(CmsFormViewBase):
    MCLS = SiteSettingTable
    title = '系统设置'
    add_url_rules = [['/setting/', 'setting']]
    template = 'fenXi/setting.html'

    def encry_password(self, raspwd):
        return generate_password_hash(raspwd)

    def view_get(self):
        if 'systemManage' not in self.current_admin_user.permissions and not self.current_admin_user.is_superadmin:
            return abort(404)
        self.context['title'] = self.title
        site_config = self.MCLS.find_one() or {}
        self.context['site_config'] = site_config
        ftp_data = FtpTable.find_one({}) or {}
        self.context['ftp_data'] = ftp_data
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'post_site_data':
            ftp_host = self.request_data.get('ftp_host') or ''
            ftp_username = self.request_data.get('ftp_username') or ''
            ftp_password = self.request_data.get('ftp_password') or ''
            server_path = self.request_data.get('server_path') or ''
            blacklistIp = self.request_data.get('blacklistIp')
            restore_data_pwd = self.request_data.get('restore_data_pwd')

            site_config = self.MCLS.find_one() or {}
            site_config['blacklistIp'] = blacklistIp or ''
            if restore_data_pwd:
                site_config['restore_data_pwd'] = self.encry_password(restore_data_pwd)

            self.MCLS.save(site_config)
            self.MCLS.update_site_config()

            _ftp_data = FtpTable.find_one({}) or {}
            _ftp_data.update({
                'ftp_host': ftp_host or '',
                'ftp_username': ftp_username or '',
                'ftp_password': ftp_password or '',
                'server_path': server_path or '',
            })
            FtpTable.save(_ftp_data)

            return self.xtjson.json_result()
        if self.action == 'check_link':
            ftp_host = self.request_data.get('ftp_host') or ''
            ftp_username = self.request_data.get('ftp_username') or ''
            ftp_password = self.request_data.get('ftp_password') or ''
            server_path = self.request_data.get('server_path') or ''

            if not ftp_host or not ftp_username or not ftp_password or not server_path:
                return self.xtjson.json_params_error('请先完善ftp连接信息！')

            ftpCls = FtpCls(
                host=ftp_host,
                user=ftp_username,
                pwd=ftp_password,
            )
            statu, res = ftpCls.check_link(server_path)
            if not statu:
                return self.xtjson.json_params_error(res)

            return self.xtjson.json_result(message='连接成功！')



class AdminUserManage(CmsFormViewBase):
    MCLS = CmsUserModel
    title = '员工管理'
    add_url_rules = [['/adminManage/', 'adminManage']]
    template = 'fenXi/adminManage.html'
    per_page = 20
    sort = [['_create_time', -1]]

    def format_datetime(self, data):
        if isinstance(data, datetime.datetime):
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    def _edit_form_html(self):
        if not self.MCLS.edit_field_sort():
            return self.xtjson.json_params_error('无处理方式!')

        html = f'''
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>用户名：</span></div>
                    <input type="text" class="form-control " id="username" placeholder="用户名" value="{self.data_dict.get('username')}">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>登录账户名：</span></div>
                    <input type="text" class="form-control " id="login_account" placeholder="登录账户名" value="{self.data_dict.get('login_account')}">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">zalo：</span></div>
                    <input type="text" class="form-control " id="zalo" placeholder="zalo" value="{self.data_dict.get('zalo') or ''}">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">邮箱：</span></div>
                    <input type="text" class="form-control" id="email" placeholder="邮箱" value="{self.data_dict.get('email') or ''}">
                </div>
                <div class="input-group mb-3">
                    <div class="input-group-prepend"><span class="input-group-text">备注：</span></div>
                    <input type="text" class="form-control " id="note" placeholder="备注" value="{self.data_dict.get('note') or ''}">
                </div>
                <span class="btn btn-primary swal2-styled" onclick="editChuFunc('{self.data_uuid}')" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>        
            '''
        return self.xtjson.json_result(message=html)

    def view_get(self):
        self.context['title'] = self.title
        self.context['add_field_sort'] = self.MCLS.add_field_sort()
        self.context['edit_field_sort'] = self.MCLS.edit_field_sort()
        self.context['format_datetime'] = self.format_datetime
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == '_add_form_html':
            res = self.checkPermission('adminManage_add')
            if res:
                return res
            if not self.MCLS.add_field_sort():
                return self.xtjson.json_params_error('无处理方式!')
            html = '''
                    <div class="input-group mb-3">
                        <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>用户名：</span>
                        </div>
                        <input type="text" class="form-control " id="username" placeholder="用户名"></div>
                    <div class="input-group mb-3">
                        <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>登录账户名：</span>
                        </div>
                        <input type="text" class="form-control " id="login_account" placeholder="登录账户名"></div>
                    <div class="input-group mb-3">
                        <div class="input-group-prepend"><span class="input-group-text"><span class="text-danger">*</span>密码：</span>
                        </div>
                        <input type="text" class="form-control " id="password" placeholder="密码"></div>
                    <div class="input-group mb-3">
                        <div class="input-group-prepend"><span class="input-group-text">zalo：</span></div>
                        <input type="text" class="form-control " id="zalo" placeholder="zalo"></div>
                    <div class="input-group mb-3">
                        <div class="input-group-prepend"><span class="input-group-text">邮箱：</span></div>
                        <input type="text" class="form-control " id="email" placeholder="邮箱"></div>
                    <div class="input-group mb-3">
                        <div class="input-group-prepend"><span class="input-group-text">备注：</span></div>
                        <input type="text" class="form-control " id="note" placeholder="备注"></div>

                    <span class="btn btn-primary swal2-styled" onclick="addChuFunc()" style="margin: 20px 10px;">确定</span>
                    <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>        
                '''
            return self.xtjson.json_result(message=html)
        if self.action == '_add_form_data':
            res = self.checkPermission('adminManage_add')
            if res:
                return res
            for db_field in self.MCLS.add_field_sort():
                _v = self.request_data.get(db_field)
                if db_field == 'password':
                    _v = self.MCLS.encry_password(_v)
                self.data_from[db_field] = _v or ''
            self.data_from['statu'] = True
            self.data_from['permissions'] = []
            self.MCLS.insert_one(self.data_from)
            return self.xtjson.json_result()

    def get_permission(self, user_cls):
        html = '<div id="permission_div">'
        html += '<table class="table table-bordered table-hover text-center">'
        html += '<thead class="thead-light"><tr>'
        html += '<th>菜单名</th>'
        html += '<th>拥有权重</th>'
        html += '</tr></thead>'
        html += '<tbody>'
        if self.check_superdamin(self.data_dict.get('permissions')):
            html += '<tr>'
            html += '<td>最高权限</td>'
            html += '<td align="left"><span class="btn btn-success btn-xs"><i class="bi-check-circle-fill"></i>&ensp;superAdmin</span></td>'
            html += '</tr>'
        for k, v in PermissionCls.name_dict.items():
            if k == 'ROOT':
                continue
            html += '<tr>'
            html += '<td>%s</td>' % k
            html += '<td align="left">'
            for per, ptext in v.items():
                if per in self.data_dict.get('permissions') and not user_cls.is_superadmin:
                    html += f"""<span class="btn btn-success btn-xs" onclick="update_permission('{self.data_dict.get('uuid')}', '{per}')"><i class="bi-check-circle-fill"></i>&ensp;{ptext}</span> """
                else:
                    html += f"""<span class="btn btn-default btn-xs" onclick="update_permission('{self.data_dict.get('uuid')}', '{per}')">{ptext}</span> """
            html += '</td></tr>'
        html += '</tbody>'
        html += '</table></div>'
        return self.xtjson.json_result(message=html)

    def post_data_other_way(self):
        crrUser = self.MCLS.query_one({'uuid': self.data_uuid})
        if self.action == 'updateStatu':
            res = self.checkPermission('adminManage_edit')
            if res:
                return res
            if crrUser.is_superadmin:
                return self.xtjson.json_params_error('超级管理员状态不可切换！')
            if self.data_dict.get('statu'):
                self.data_from['statu'] = False
            else:
                self.data_from['statu'] = True
            self.data_dict.update(self.data_from)
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()
        if self.action == 'del':
            res = self.checkPermission('adminManage_del')
            if res:
                return res
            if crrUser.is_superadmin:
                return self.xtjson.json_params_error('超级管理员不可删除！')
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()
        if self.action == '_edit_form_html':
            res = self.checkPermission('adminManage_edit')
            if res:
                return res
            return self._edit_form_html()
        if self.action == '_edit_pwd_html':
            res = self.checkPermission('adminManage_edit')
            if res:
                return res
            html = f"""
                <div class="input-group mb-3">
                    <div class="input-group-prepend">
                        <span class="input-group-text">新密码：</span>
                    </div>
                    <input type="text" class="form-control" id="new_password" placeholder="输入新密码">
                </div>         
                <div class="input-group mb-3">
                    <div class="input-group-prepend">
                        <span class="input-group-text">确认密码：</span>
                    </div>
                    <input type="text" class="form-control" id="confirm_password" placeholder="输入确认密码">
                </div>
                <span class="btn btn-primary swal2-styled" onclick="updatePwdFunc('{self.data_uuid}')" style="margin: 20px 10px;">确定</span>
                <span class="btn btn-default swal2-styled" onclick="xtalert.close()" style="margin: 20px 10px;">取消</span>                        
            """
            return self.xtjson.json_result(message=html)
        if self.action == 'update_pwd':
            new_password = self.request_data.get('new_password')
            _v = self.MCLS.encry_password(new_password)
            self.data_dict['password'] = _v
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()
        if self.action == 'get_permission_html':
            return self.get_permission(crrUser)
        if self.action == '_edit_permission':
            if crrUser.is_superadmin:
                return self.xtjson.json_params_error('最高管理员权限不可修改!')
            per = self.request_data.get('p')
            if not per:
                return self.xtjson.json_params_error('操作错误!')
            u_permissions = self.data_dict.get('permissions')
            if per in u_permissions:
                u_permissions.remove(per)
            else:
                u_permissions.append(per)
            self.data_dict['permissions'] = u_permissions
            self.MCLS.save(self.data_dict)
            return self.get_permission(crrUser)



class PersonalInfoView(CmsFormViewBase):
    MCLS = CmsUserModel
    title = '个人中心'
    add_url_rules = [['/personalInfo/<string:data_uuid>', 'personalInfo']]
    template = 'fenXi/personalInfo.html'

    def format_datetime(self, data):
        if isinstance(data, datetime.datetime):
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data

    def view_get(self, data_uuid):
        user_data = self.MCLS.find_one({'uuid': data_uuid})
        if not user_data:
            return abort(404)
        self.context['title'] = self.title
        self.context['user_data'] = user_data
        self.context['format_datetime'] = self.format_datetime
        return render_template(self.template, **self.context)

    def post_data_other_way(self):
        if self.action == '_edit_user_data':
            for db_field in self.MCLS.edit_field_sort():
                field_cls = self.MCLS.fields().get(db_field)
                v = self.request_data.get(db_field)
                statu, res = field_cls.validate(v)
                if not statu:
                    return self.xtjson.json_params_error(res)
                self.data_from[db_field] = res
            self.data_dict.update(self.data_from)
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()
        if self.action == '_edit_user_pwd':
            original_password = self.request_data.get('original_password')
            new_password = self.request_data.get('new_password')
            if not new_password:
                return self.xtjson.json_params_error()
            if not self.MCLS.check_password(self.data_dict.get('password'), original_password.strip()):
                return self.xtjson.json_params_error('原密码错误!')
            self.data_from['password'] = self.MCLS.encry_password(new_password.strip())
            self.data_dict.update(self.data_from)
            self.MCLS.save(self.data_dict)
            return self.xtjson.json_result()



class BettingDataView(CmsFormViewBase):
    MCLS = BettingDataTable
    title = '输赢分层'
    add_url_rules = [['/bettingSy/', 'bettingSy']]
    template = 'fenXi/bettingSy.html'
    per_page = 10
    sort = [['create_time', -1]]

    def format_money(self, data):
        try:
            return format(int(data), ",")
        except:
            return data

    def fenXiCsv_func(self, filePath):
        ''' 分析数据 '''
        name_dict = {
            '账号': 'account',
            '上级代理': 'parent_agent',
            '下注笔数': 'betting_count',
            '投注金额': 'betting_money',
            '盈利投注（金额）': 'profit_money',
            '代理赔率（金额）': 'agent_odds',
            '代理返水（金额）': 'agent_backwater',
            '会员输赢（不含退水）': 'vip_winlose_money',
            '实际退水': 'actual_backwater',
            '实际输赢（含退水）': 'actual_winlose_money',
        }

        try:
            with open(filePath, 'r', encoding='gbk') as fp:
                csv_data = list(csv.DictReader(fp))
        except:
            try:
                with open(filePath, 'r', encoding='utf8') as fp:
                    csv_data = list(csv.DictReader(fp))
            except:
                with open(filePath, 'r', encoding='GB2312') as fp:
                    csv_data = list(csv.DictReader(fp))

        batch_code = int(time.time() * 1000)
        pool = multiprocessing.Pool(15)
        multi_result = []
        while csv_data:
            crr_datas = csv_data[:100]
            if not crr_datas:
                break
            del csv_data[:100]
            multi_result.append(pool.apply_async(func=fcFenXiCsv_fucn, args=(crr_datas, name_dict, batch_code)))
        pool.close()
        pool.join()
        zd_fenxi_func(csv_data, name_dict)

        total_datas = []
        for m in multi_result:
            total_datas += m.get(0)

        return total_datas

    def blacklist_func(self, datas, dataKey):
        ''' 加入黑名单 '''
        for d in datas:
            self.MCLS.update_one({'account': d},{'$set':{'blacklistState': True}})
            SiteRedis.incrby(dataKey)

    def toBlacklist_func(self, datas, dataKey):
        ''' 解除黑名单 '''
        for d in datas:
            self.MCLS.update_one({'account': d},{'$set':{'blacklistState': False}})
            SiteRedis.incrby(dataKey)

    def import_data(self, datas, dataKey):
        ''' 导入数据 '''
        for da in datas:
            self.MCLS.update_one({'account': da.get('account')}, {'$set': da},upsert=True)
            SiteRedis.incrby(dataKey)

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['账户', '实际输赢（含退水）']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            for data in datas:
                row += 1
                wa.cell(row=row, column=1, value=str(data.get('account') or ''))
                wa.cell(row=row, column=2, value=str( self.format_money(data.get('actual_winlose_money') or '0')) )

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def del_all(self):
        self.MCLS.delete_many({})

    def view_get(self):
        if 'bettingSy' not in self.current_admin_user.permissions and not self.current_admin_user.is_superadmin:
            return abort(404)

        self.context['title'] = self.title
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['format_money'] = self.format_money
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'importData':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('csv'):
                return self.xtjson.json_params_error('文件格式错误，只支持CSV文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/importFile/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)

            datas = self.fenXiCsv_func(import_folder + new_filename + fext)
            if not datas:
                return self.xtjson.json_params_error('文件无数据！')

            dataKey = 'progress_' + str(int(time.time()*100))
            SiteRedis.incrby(dataKey, 1)
            SiteRedis.expire(dataKey, 60*20)
            threading.Thread(target=self.import_data, args=(datas, dataKey)).start()
            _data = {
                'dataKey': dataKey,
                'total': len(datas),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'del_all':
            threading.Thread(target=self.del_all).start()
            return self.xtjson.json_result(message='删除中， 请稍后刷新查看！')
        if self.action == 'importBlacklistState':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('csv'):
                return self.xtjson.json_params_error('文件格式错误，只支持CSV文件上传！')

            dataf = fileobj.read().decode()

            datas = []
            for d in dataf.split('\n')[1:]:
                _v = d.strip().strip(',')
                if _v:
                    datas.append(_v)

            dataKey = 'progress_' + str(int(time.time()*100))
            SiteRedis.incrby(dataKey, 1)
            SiteRedis.expire(dataKey, 60*20)
            threading.Thread(target=self.blacklist_func, args=(datas, dataKey)).start()
            _data = {
                'dataKey': dataKey,
                'total': len(datas),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'jcBlacklistState':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('csv'):
                return self.xtjson.json_params_error('文件格式错误，只支持CSV文件上传！')

            dataf = fileobj.read().decode()

            datas = []
            for d in dataf.split('\n')[1:]:
                _v = d.strip().strip(',')
                if _v:
                    datas.append(_v)

            dataKey = 'progress_' + str(int(time.time()*100))
            SiteRedis.incrby(dataKey, 1)
            SiteRedis.expire(dataKey, 60*20)
            threading.Thread(target=self.toBlacklist_func, args=(datas, dataKey)).start()
            _data = {
                'dataKey': dataKey,
                'total': len(datas),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'getProgres':
            dataKey = self.request_data.get('dataKey')
            if not dataKey:
                return self.xtjson.json_params_error()
            _v = int(SiteRedis.get(dataKey).decode())
            return self.xtjson.json_result(data={'v': _v})
        if self.action == 'exportCrrData':
            filter_dict = {}
            _crd = self.MCLS.find_one({'blacklistState': False}, sort=[['batch_code', -1]]) or {}
            if _crd:
                filter_dict['batch_code'] = _crd.get('batch_code')

            datas = self.MCLS.find_many(filter_dict, sort=[['actual_winlose_money', -1]])

            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '输赢分层，当次输赢信息-导出数据' + datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)

            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到 导出数据 下载！')

    def post_data_other_way(self):
        if self.action == 'del':
            self.MCLS.delete_one({'uuid': self.data_uuid})
            return self.xtjson.json_result()



class layeredResultView(CmsFormViewBase):
    MCLS = BettingDataTable
    title = '输赢分层-分析结果'
    add_url_rules = [['/layeredResult/', 'layeredResult']]
    template = 'fenXi/layeredResult.html'
    per_page = 10
    sort = [['create_time', -1]]

    def format_money(self, data):
        try:
            return format(int(data), ",")
        except:
            return data

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['账户', '原来层级', '更换层级', '实际输赢（含退水）']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            for data in datas:
                row += 1
                wa.cell(row=row, column=1, value=str(data.get('account') or ''))
                wa.cell(row=row, column=2, value=str( self.format_money(data.get('low_quota') or '')) )
                wa.cell(row=row, column=3, value=str( self.format_money(data.get('new_quota') or '')) )
                wa.cell(row=row, column=4, value=str( self.format_money(data.get('actual_winlose_money') or '')) )

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def view_get(self):
        self.context['title'] = self.title
        self.context['format_money'] = self.format_money

        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res
        filter_dict.update(res[0])
        context_res.update(res[1])
        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())

        filter_dict['blacklistState'] = False
        filter_dict['promotion_state'] = True
        _crd = self.MCLS.find_one({'blacklistState': False}, sort=[['batch_code', -1]]) or {}
        if _crd:
            filter_dict['batch_code'] = _crd.get('batch_code')

        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'exportData':
            filter_dict = {}
            filter_dict['blacklistState'] = False
            filter_dict['promotion_state'] = True
            _crd = self.MCLS.find_one({'blacklistState': False}, sort=[['batch_code', -1]]) or {}
            if _crd:
                filter_dict['batch_code'] = _crd.get('batch_code')
            datas = self.MCLS.find_many(filter_dict, sort=self.sort)

            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '输赢分层-导出数据'
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)

            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到 导出数据 下载！')



class ChongZhiFenXiCls(CmsFormViewBase):
    title = '充值分析'
    show_menu = False
    add_url_rules = [['/czFenxi/', 'czFenxi']]
    template = 'fenXi/chognzhi_fenxi.html'
    MCLS = ChongZhiFenXiTable

    def readXlsx_data(self, filepath):
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        wa = wb.active
        _crrdata = {}
        cll = 0
        for row in wa.rows:
            cll += 1
            if cll == 1:
                continue
            _d = []
            for c in row:
                _d.append(c.value)
            _crrdata[str(_d[0]).replace(' ','').replace('=','').replace('"', '')] = int(_d[2] or 0)
        wb.close()
        return _crrdata

    def fenxi_func(self, crrdata1, crrdata2, dataKey):
        self.MCLS.delete_many({})
        for k,v in crrdata1.items():
            SiteRedis.incrby(dataKey)
            _vv = crrdata2.get(k) or 0
            _d = {
                'account': k,
                'money': float(v - _vv)
            }
            self.MCLS.insert_one(_d)

    def view_get(self):
        self.context['title'] = self.title
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'importData1':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('xlsx'):
                return self.xtjson.json_params_error('文件格式错误，只支持xlsx文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/importFenxiFile/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)

            return self.xtjson.json_result(message=(import_folder + new_filename + fext).replace(current_app.root_path + '/' + self.project_static_folder , ''))

        if self.action == 'importData2':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('xlsx'):
                return self.xtjson.json_params_error('文件格式错误，只支持xlsx文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/importFenxiFile/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            new_filename = shortuuid.uuid()
            fileobj.save(import_folder + new_filename + fext)

            return self.xtjson.json_result(message=(import_folder + new_filename + fext).replace(current_app.root_path + '/' + self.project_static_folder , ''))

        if self.action == 'fenxiData':
            filepath1 = self.request_data.get('filepath1')
            filepath2 = self.request_data.get('filepath2')
            if not filepath1 or not filepath2:
                return self.xtjson.json_params_error()

            crrdata1 = self.readXlsx_data(current_app.root_path + '/' + self.project_static_folder + filepath1)
            crrdata2 = self.readXlsx_data(current_app.root_path + '/' + self.project_static_folder + filepath2)

            dataKey = 'progress_' + str(int(time.time()*100))
            SiteRedis.incrby(dataKey, 1)
            SiteRedis.expire(dataKey, 60*20)
            threading.Thread(target=self.fenxi_func, args=(crrdata1, crrdata2, dataKey)).start()
            _data = {
                'dataKey': dataKey,
                'total': len(crrdata1),
            }
            return self.xtjson.json_result(data=_data)
        if self.action == 'getProgres':
            dataKey = self.request_data.get('dataKey')
            if not dataKey:
                return self.xtjson.json_params_error()
            _v = int(SiteRedis.get(dataKey).decode())
            return self.xtjson.json_result(data={'v': _v})



class czfxResultCls(CmsFormViewBase):
    title = '充值分析结果'
    show_menu = False
    add_url_rules = [['/czfxResult/', 'czfxResult']]
    template = 'fenXi/czfx_result.html'
    MCLS = ChongZhiFenXiTable
    per_page = 50
    sort = [['money', -1]]

    def exportData(self, datas, log_uuid, export_folder, filename):
        ''' 导出数据 '''
        export_data = ExportDataModel.find_one({'uuid': log_uuid})
        if not export_data:
            return
        try:
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            crr_count = 0

            wb = Workbook()
            wa = wb.active
            row = 1
            header = ['账户', '金额差值']
            for h in range(len(header)):
                wa.cell(row=row, column=h+1, value=header[h])

            for data in datas:
                row += 1
                wa.cell(row=row, column=1, value=str(data.get('account') or ''))
                wa.cell(row=row, column=2, value=str(data.get('money') or '0'))

                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def view_get(self):
        self.context['title'] = self.title
        page = request.args.get('page', 1, int)
        skip = (page - 1) * self.per_page
        self.context['title'] = self.title
        filter_dict, context_res = {}, {}
        fields = self.MCLS.fields()
        self.context['FIELDS'] = fields
        statu, res = self.search_func(fields)
        if not statu:
            return res

        filter_dict.update(res[0])
        context_res.update(res[1])

        self.context.update(self.get_context())
        filter_dict.update(self.get_filter_dict())
        total = self.MCLS.count(filter_dict)
        all_datas = self.MCLS.find_many(filter_dict, limit=self.per_page, skip=skip, sort=self.sort)

        pagination = PagingCLS.pagination(page, self.per_page, total)
        self.context['total'] = total
        self.context['all_datas'] = all_datas
        self.context['pagination'] = pagination
        self.context['search_res'] = context_res
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'daoChuJieGuo':
            datas = self.MCLS.find_many({}, sort=[['money', -1]])

            absolute_folter = os.path.join(current_app.root_path, self.project_static_folder)
            export_folder = os.path.join(absolute_folter, ASSETS_FOLDER, EXPORT_FOLDER)
            filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S_') + str(random.choice(range(100, 999))) + '.xlsx'
            _out_data_dict = {
                'filename': filename,
                'statu': ExportStatu.ongoing,
                'path': os.path.join(export_folder, filename).replace(absolute_folter, ''),
                'total': len(datas),
                'out_count': 0,
                'note': '充值分析-导出数据' + datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            }
            uuid = ExportDataModel.insert_one(_out_data_dict)

            threading.Thread(target=self.exportData, args=(datas, uuid, export_folder, filename)).start()
            return self.xtjson.json_result(message='数据导出中，请稍后到 导出数据 下载！')



class CustomerAnalysisCls(CmsFormViewBase):
    title = '客服注单分析'
    show_menu = False
    add_url_rules = [['/customerAnalysis/', 'customerAnalysis']]
    template = 'fenXi/customerAnalysis.html'

    def fenXi_csv(self, filePath):
        name_dict = {
            '订单号': 'order_number',
            '会员': 'member_account',
            '下注时间': 'create_time',
            '游戏': 'game_name',
            '期号': 'date_number',
            '玩法分类': 'play_type',
            '玩法名称': 'play_name',
            '注单信息': 'zhudan_info',
            '下注金额': 'money',
            '会员输赢': 'win_lose',
        }

        csv_data = []
        try:
            with open(filePath, 'r', encoding='gbk') as fp:
                csv_data = list(csv.DictReader(fp))
        except:
            try:
                with open(filePath, 'r', encoding='utf8') as fp:
                    csv_data = list(csv.DictReader(fp))
            except:
                with open(filePath, 'r', encoding='GB2312') as fp:
                    csv_data = list(csv.DictReader(fp))

        pool = multiprocessing.Pool(15)
        multi_result = []
        while csv_data:
            crr_datas = csv_data[:5000]
            if not crr_datas:
                break
            del csv_data[:5000]
            multi_result.append(pool.apply_async(func=zd_fenxi_func, args=(crr_datas, name_dict)))
        pool.close()
        pool.join()
        zd_fenxi_func(csv_data, name_dict)

        total_datas = []
        for m in multi_result:
            total_datas += m.get(0)

        return total_datas

    def view_get(self):
        self.context['title'] = self.title

        _ppimport_folder = current_app.root_path + '/' + self.project_static_folder + '/assets/importFenxiFile/' + datetime.datetime.now().strftime('%Y%m%d') + '/'
        path_ls = []
        if os.path.exists(_ppimport_folder):
            for p in os.listdir(_ppimport_folder):
                _p = (_ppimport_folder + p).replace(current_app.root_path + '/' + self.project_static_folder , '')
                _dit = {
                    'path': _p,
                    'file_name': p,
                }
                path_ls.append(_dit)

        self.context['path_ls'] = path_ls
        return render_template(self.template, **self.context)

    def post_other_way(self):
        if self.action == 'importData1':
            fileobj = request.files['upload']
            fname, fext = os.path.splitext(fileobj.filename)
            if not fext.endswith('csv'):
                return self.xtjson.json_params_error('文件格式错误，只支持csv文件上传！')
            import_folder = current_app.root_path + '/' + self.project_static_folder + '/assets/importFenxiFile/' + datetime.datetime.now().strftime('%Y%m%d') + '/'
            if not os.path.exists(import_folder):
                os.makedirs(import_folder)

            # new_filename = shortuuid.uuid()
            new_filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            fileobj.save(import_folder + new_filename + fext)
            file_path = (import_folder + new_filename + fext).replace(current_app.root_path + '/' + self.project_static_folder , '')
            _data = {
                'file_path': file_path,
                'filename': new_filename + fext
            }
            return self.xtjson.json_result(data=_data, message=file_path)
        if self.action == 'fenxiData':
            filepath1 = self.request_data.get('filepath1')
            CustomerName = self.request_data.get('CustomerName')
            if not filepath1 or not CustomerName:
                return self.xtjson.json_params_error('分析失败！')
            _p = current_app.root_path + '/' + self.project_static_folder + filepath1
            if not os.path.exists(_p):
                return self.xtjson.json_params_error('分析失败！')

            datas = self.fenXi_csv(_p)
            wf_fx_dict = {}
            for da in datas:
                play_name = da.get('play_name') or ''
                if not play_name:
                    continue
                member_account = da.get('member_account') or ''
                if member_account != CustomerName.strip():
                    continue
                zhudan_info = da.get('zhudan_info') or ''
                zhudan_info = zhudan_info.strip().split(' ')

                _fx_ls = get_new_zhudan_info_func(zhudan_info, play_name) or []
                _dls = wf_fx_dict.get(play_name) or []
                _dls += _fx_ls
                wf_fx_dict.update({play_name: _dls})

            __wf_fx_dict = {}
            for k, v in wf_fx_dict.items():
                __wf_fx_dict.update({k: ','.join(sorted(list(set(v))))})
            numberFxHtml = '''
            <li class="list-group-item mt-3">
                <table class="table table-bordered" style="width: 80%; margin: auto;">
                    <tbody>
                        <tr>
                            <td>玩法名称</td>
                            <td>下注号码</td>
                            <td>号码数量</td>                            
                        </tr>                    
            '''
            for k, v in __wf_fx_dict.items():
                numberFxHtml += f'''
                        <tr>
                            <td>{k}</td>
                            <td>{v}</td>
                            <td>{len(v.split(','))}</td>
                        </tr>     
                '''
            numberFxHtml += '</tbody></table></li>'

            html = f'''
            <div class="panel">
                <h3>
                    <svg style="color: #1E9FFF;" viewBox="64 64 896 896" focusable="false" data-icon="hourglass" width="20" height="20" fill="currentColor" aria-hidden="true"><path d="M742 318V184h86c4.4 0 8-3.6 8-8v-56c0-4.4-3.6-8-8-8H196c-4.4 0-8 3.6-8 8v56c0 4.4 3.6 8 8 8h86v134c0 81.5 42.4 153.2 106.4 194-64 40.8-106.4 112.5-106.4 194v134h-86c-4.4 0-8 3.6-8 8v56c0 4.4 3.6 8 8 8h632c4.4 0 8-3.6 8-8v-56c0-4.4-3.6-8-8-8h-86V706c0-81.5-42.4-153.2-106.4-194 64-40.8 106.4-112.5 106.4-194zm-72 388v134H354V706c0-42.2 16.4-81.9 46.3-111.7C430.1 564.4 469.8 548 512 548s81.9 16.4 111.7 46.3C653.6 624.1 670 663.8 670 706zm0-388c0 42.2-16.4 81.9-46.3 111.7C593.9 459.6 554.2 476 512 476s-81.9-16.4-111.7-46.3A156.63 156.63 0 01354 318V184h316v134z"></path></svg>
                    【<b>{CustomerName}</b>】：数据分析结果
                </h3>
                <div class="panel-body">
                    <ul class="list-group">
                        <div id="wanFanTop">
                            {numberFxHtml}
                        </div>
                    </ul>
                </div>
            </div>        
            '''
            return self.xtjson.json_result(data={'html': html})
        if self.action == 'delfile':
            filepath = self.request_data.get('filepath')
            if not filepath:
                return self.xtjson.json_params_error()
            _p = current_app.root_path + '/' + self.project_static_folder + filepath
            cmd = 'rm ' + _p
            os.popen(cmd)
            return self.xtjson.json_result(message='删除成功！')


